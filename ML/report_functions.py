# Importing the required libraries
import os
import pandas as pd
import numpy as np
import regex as re
import spacy
nlp = spacy.load("en_core_web_sm")
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill,Border,Side
from openpyxl.drawing.text import Font as Font_chart
from openpyxl.chart import BarChart, Reference,Series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.marker import Marker
# Need to change acc to the user's path
unnecessary_keywords_file_path = os.path.join(os.path.dirname(__file__), "unnecessary_keywords.txt")
unnecessary_keywords = np.loadtxt(unnecessary_keywords_file_path, dtype=str, delimiter=" ")
from thefuzz import fuzz, process
import json
from spacy.training import Example
# Need to change acc to the user's path
ner_model_file_path = os.path.join(os.path.dirname(__file__), "ner_model")
nlp = spacy.load(ner_model_file_path)
# Function to train the NER model on the description column of the dataframe
def ner_model_trainer(info_extracted_df,iterations = 20):
    inputs = info_extracted_df['Description'].values.tolist()
    entities = info_extracted_df['Cleaned Entity Name'].values.tolist()

    training_data = []
    for i in range(len(inputs)):
        entity = entities[i]
        input_text = inputs[i]
        if entity != "Not Found" and entity != "":
            start = input_text.find(entity)
            end = start + len(entity)
            training_data.append((input_text, {"entities": [(start, end, "ENTITY")]}))
    # Need to change acc to the user's path
    nlp = spacy.load(os.path.join(os.getcwd(),"ner_model"))
    ner = nlp.get_pipe("ner")
    other_pipes = [pipe for pipe in nlp.pipe_names if pipe != "ner"]

    with nlp.disable_pipes(*other_pipes):
        optimizer = nlp.create_optimizer()
        for itn in range(iterations):
            losses = {}
            examples = training_data
            for text, annotations in examples:
                doc = nlp.make_doc(text)
                example = Example.from_dict(doc, annotations)
                nlp.update([example], drop=0.5, sgd=optimizer, losses=losses)
    # Need to change acc to the user's path

    output_dir = os.path.join(os.path.dirname(__file__), "ner_model")
    nlp.to_disk(output_dir)

# Function to group the entity names
def entity_name_grouper(name_list):
    groups  = {}
    group_num = 0
    for i in range(len(name_list)):
        # Travel through the groups dictionary and find the group which has maximum value of minimum similarity withi the group
        max_similarity = 0
        max_group = -1

        for key, value in groups.items():
            # Find the minimum similarity of the name with the group
            min_similarity = max([fuzz.token_set_ratio(name_list[i], name) for name in value])
            if min_similarity > max_similarity:
                max_similarity = min_similarity
                max_group = key

        if max_similarity >= 80:
                groups[max_group].append(name_list[i])
        else :
            # Create a new group
            groups[group_num] = [name_list[i]]
            group_num += 1
    return groups

# Main function to consolidate the final excels of all the pdfs/excels into a single excel
def sheet_consolidator(sheets_path_list):
    combined_sheets = []

    for sheet in sheets_path_list:
        sheet = pd.read_csv(sheet)
        sheet['Date'] = pd.to_datetime(sheet['Date'], errors='coerce')
        sheet = sheet.dropna(subset=['Debit', 'Credit'], how='all')
        sheet['Description'] = sheet['Description'].astype(str)
        sheet['Description'] = sheet['Description'].str.lower()
        sheet['Description'] = sheet['Description'].str.replace('[^a-zA-Z0-9\s]', ' ', regex=True)

        sheet['Debit'] = pd.to_numeric(sheet['Debit'], errors='coerce')
        sheet['Credit'] = pd.to_numeric(sheet['Credit'], errors='coerce')
        sheet['Balance'] = pd.to_numeric(sheet['Balance'], errors='coerce')
        sheet = sheet.dropna(subset=['Balance'], how='all')
        sheet = sheet.reset_index(drop=True)
        combined_sheets.append(sheet) # type: ignore

    combined_sheets = pd.concat(combined_sheets, ignore_index=True)
    combined_sheets = combined_sheets.sort_values(by='Date', ascending=True)
    combined_sheets = combined_sheets.drop_duplicates()
    combined_sheets = combined_sheets.reset_index(drop=True)

    combined_sheets['Year'] = combined_sheets['Date'].dt.year
    combined_sheets['Quarter'] = combined_sheets['Date'].dt.quarter
    combined_sheets['Month'] = combined_sheets['Date'].dt.month
    combined_sheets['Month Year'] = combined_sheets['Date'].dt.strftime('%b %Y')
    combined_sheets['Day Name'] = combined_sheets['Date'].dt.day_name()
    combined_sheets['Date'] = combined_sheets['Date'].dt.strftime('%d-%m-%Y')

    return combined_sheets

# Function to remove duplicate words from a string
def remove_duplicates(input_string):
    return ' '.join(dict.fromkeys(input_string.split()))

# Function to process the details sheet
def sheet_details_processor(details):
    details['Statement Start'] = pd.to_datetime(details['Statement Start'], errors='coerce')
    details['Statement End'] = pd.to_datetime(details['Statement End'], errors='coerce')
    from_date = details['Statement Start'].min()
    to_date = details['Statement End'].max()
    from_date = from_date.strftime('%d-%b-%Y')
    to_date = to_date.strftime('%d-%b-%Y')
    ifsc_list = details['IFSC'].dropna().unique().tolist()
    bank_list = details['Bank'].dropna().unique().tolist()
    account_type_list = details['Account Type'].dropna().unique().tolist()

    return from_date, to_date, ifsc_list, bank_list, account_type_list

# Transaction Type keywords
transaction_keywords = {
    "charges" : ["charges", "charge", "service charge", "commission","fee","chrg","chgs","chg"],
    "cheque": ["by cheque","cheque", "chq", "chq no", "cheque no", "clearing", "chq deposit","cts","micr"],
    "cash": ["by cash", "cash deposit", "cash withdrawal", "cdm", "atm","cash","cash wdl"],
    "internet banking": ["internet banking", "net banking", "online banking", "netbanking", "inb", "internet transfer"],
    "rtgs": ["rtgs", "rtgs inward", "rtgs outward", "rtgs transfer"],
    "neft": ["neft", "neft transfer", "neft inward", "neft outward"],
    "upi": ["upi", "upi transfer", "upi payment", "upi id", "vpa", "phonepe", "google pay", "gpay", "paytm", "bhim"],
    "edc": ["edc", "edc debit", "edc credit"],
    "ecs": ["ecs", "ecs debit", "ecs credit"],
    "imps": ["imps", "imps transfer", "imps inward", "imps outward"],
    "debit card/credit card": ["debit card", "credit card","e-commerce", "online payment","card"],
    "demand draft": ["dd", "demand draft", "dd number"],
    "wire transfer": ["wire", "swift", "tt"],
    "cms" : ["cms"],
    "transfer": ["transfer","trf", "tnf", "fund transfer", "transfer to", "transfer from","ft"],
    "clearing house transaction" : ["clg","clearing house"],
    "original credit transaction" : ["oct", "octroi"],
    "brought forward" : ["b/f", "brought forward","bf","opening balance"],
    "cin" : ["cin"]
}

# List of online methods
online_methods = ["upi","rtgs","neft","internet banking","imps"]

# Purpose keywords
purpose_keywords = {
    "loan" : ["emi", "loan", "interest", "principal", "emi number", "emi date", "emi amount", "emi payment", "loan number", "loan amount", "loan payment", "loan interest", "loan principal"],
    "salary" : ["salary", "wages"],
    "rent" : ["rent"],
    "gst tax" : ["gst", "gst payment"],
    "tax" : ["tax", "tds", "income tax","tax payment"],
    "insurance" : ["insurance", "premium", "insurance payment"],
    "investment" : ["investment", "mutual fund", "shares", "stock", "investment payment"],
    "utility" : ["electricity", "water", "gas", "utility", "utility payment","recharge"],
    "credit interest" : ["credit interest", "interest credit"],
    "interest" : ["interest", "interest payment"],
    "entertainment" : ["entertainment", "movie", "theatre", "entertainment payment"],
    "charges" : ["charges", "charge", "service charge", "commission","fee"],
    "food": ["food", "restaurant", "hotel", "food payment","swiggy","zomato"],
    "shopping": ["shopping", "mall", "shop", "store", "shopping payment","amazon","flipkart"],
    "travel": ["travel", "flight", "train", "bus", "travel payment","makemytrip","goibibo"],
}

# Charges keywords
charges_keywords = {
    "return" : ["return", "return charges","rtn"],
    "atm" : ["atm", "atm withdrawal", "atm charges"],
    "debit card" : ["debit card", "debit card charges"],
    "credit card" : ["credit card", "credit card charges"],
    "card" : ["card"],
    "consolidated charges" : ["consolidated charges"],
    "gst on charges" : ["gst"],
    "rtgs charges" : ["rtgs"],
    "alerts charges" : ["alerts"],
    "netbanking charges" : ["netbanking"],
    "clearance charges" : ["clearance"],
    "imps" : ["imps"]
}

# Preprocess the description text to remove descepencies
def preprocess_text(text):
    text = text.lower()
    text = text.replace("  "," ")
    text.strip()
    words = text.split()

    # Filter out stop words and single or two-letter words
    filtered_words = [word for word in words if word not in unnecessary_keywords and len(word) > 2]

    # remove duplicate words while preserving the order
    filtered_words = list(dict.fromkeys(filtered_words))

    # Join the filtered words back into a single string
    cleaned_text = ' '.join(filtered_words)

    # remove any word containing numbers -
    cleaned_text = ' '.join([i for i in cleaned_text.split() if not any(c.isdigit() for c in i)])

    if len(cleaned_text) == 0:
        cleaned_text = "Not Found"

    return cleaned_text

# Function to use NER model to extract entities from the description
def extract_entity_name(transaction):
    # Preprocess the transaction text
    cleaned_text = preprocess_text(transaction)

    # Process the cleaned text with the NER model
    doc = nlp(cleaned_text)
    # Extract entities
    entities = []
    for ent in doc.ents:
        # Consider only PERSON and ORG (organization) entities
        if ent.label_ in ["ENTITY"]:
            entities.append(ent.text)

    # Return the extracted entities
    return entities

# Main function to extract the information from the description
def description_info_extraction(transaction_df,party_name,use_ner_model = False):
    nlp = spacy.load(ner_model_file_path)
    party_name = party_name.lower()
    # remove pvt ltd from the party name
    party_name = party_name.replace("pvt","")
    party_name = party_name.replace("ltd","")

    party_name = party_name.replace("private","")
    party_name = party_name.replace("limited","")
    # remove special characters from the party name
    party_name = re.sub('[^a-zA-Z0-9\s]', " ", party_name)
    party_name = party_name.strip()

    # Extracting the transaction type from the description
    transaction_df['Transaction Type'] = np.nan
    transaction_df['Purpose'] = np.nan
    transaction_df['Charge Type'] = np.nan
    transaction_df['Direction'] = np.nan
    transaction_df['Related Party'] = np.nan
    transaction_df['Entity'] = np.nan
    transaction_df['Cheque Return'] = np.nan
    transaction_df['Online Return'] = np.nan
    transaction_df['ECS Return'] = np.nan
    transaction_df['Cleaned Entity Name'] = np.nan
    transaction_df['Reconciled Group'] = np.nan
    transaction_df["Reconciled Name"] = np.nan
    for index, row in transaction_df.iterrows():
        row['Description'] = remove_duplicates(row['Description'])
        description = row['Description']
        # remove duplicate values from row['Description']

        # if the description only has "b f" or "bf" then it is brought forward as transaction type
        if use_ner_model:
            doc = nlp(description)
            if len(doc.ents) > 0:
                # join all the entities found in the description
                entity = ' '.join([ent.text for ent in doc.ents])
                transaction_df.at[index, 'Cleaned Entity Name'] = entity
            else:
                transaction_df.at[index, 'Cleaned Entity Name'] = "Not Found"
        else :
            transaction_df.at[index, 'Cleaned Entity Name'] = preprocess_text(description)
        if not np.isnan(row['Debit']) and row['Debit'] > 0:
            transaction_df.at[index, 'Direction'] = "outward"
        elif not np.isnan(row['Credit']) and row['Credit'] > 0:
            transaction_df.at[index, 'Direction'] = "inward"
        else :
            transaction_df.at[index, 'Direction'] = "other"

        if description.strip() == "b f" or description.strip() == "bf":
            transaction_df.at[index, 'Transaction Type'] = "brought forward"
            transaction_df.at[index, 'Purpose'] = "brought forward"
            transaction_df.at[index, 'Cleaned Entity Name'] = "brought forward"
        else:
            for transaction_type, keywords in transaction_keywords.items():
                flag = 0
                for keyword in keywords:
                    if keyword in description:
                        transaction_df.at[index, 'Transaction Type'] = transaction_type
                        flag = 1
                        break
                if flag == 1:
                    break
            if pd.isna(transaction_df.at[index, 'Transaction Type']):
                transaction_df.at[index, 'Transaction Type'] = "other"

        # if transaction type is charges, then check for the sub type of charges
        if transaction_df.at[index, 'Transaction Type'] == "charges":
            for charge_type, charge_keywords in charges_keywords.items():
                flag = 0
                for keyword in charge_keywords:
                    if keyword in description:
                        transaction_df.at[index,'Charge Type'] = charge_type
                        flag = 1
                        break
                if flag == 1:
                    break
            if pd.isna(transaction_df.at[index,'Charge Type']):
                transaction_df.at[index, 'Purpose'] = "others"
        else :
            transaction_df.at[index, 'Charge Type'] = "not applicable"

        for purpose, keywords in purpose_keywords.items():
            flag = 0
            for keyword in keywords:
                if keyword in description:
                    transaction_df.at[index, 'Purpose'] = purpose
                    flag = 1
                    break
            if flag == 1:
                break
        if pd.isna(transaction_df.at[index, 'Purpose']):
            transaction_df.at[index, 'Purpose'] = "others"

        if transaction_df.at[index, 'Transaction Type'] == "cheque":
            if "return" in description:
                transaction_df.at[index, 'Cheque Return'] = "yes"
            else:
                transaction_df.at[index, 'Cheque Return'] = "no"
        else:
            transaction_df.at[index, 'Cheque Return'] = "not applicable"

        if transaction_df.at[index, 'Transaction Type'] in online_methods:
            if "return" in description:
                transaction_df.at[index, 'Online Return'] = "yes"
            else:
                transaction_df.at[index, 'Online Return'] = "no"
        else:
            transaction_df.at[index, 'Online Return'] = "not applicable"

        if transaction_df.at[index, 'Transaction Type'] == "ecs":
            if "return" in description:
                transaction_df.at[index, 'ECS Return'] = "yes"
            else:
                transaction_df.at[index, 'ECS Return'] = "no"
        else:
            transaction_df.at[index, 'ECS Return'] = "not applicable"

        # take the description and remove the transaction type and purpose from it
        cropped_description = description.replace(transaction_df.at[index, 'Transaction Type'], "")
        cropped_description = cropped_description.replace(transaction_df.at[index, 'Purpose'], "")

        for party_keyword in party_name.split():
            if party_keyword in cropped_description:
                transaction_df.at[index, 'Related Party'] = "related party"
                transaction_df.at[index,'Purpose'] = "Inhouse Transaction"
                break
        if pd.isna(transaction_df.at[index,'Related Party']):
            transaction_df.at[index,'Related Party'] = "not related party"

        # identify entity name
        entities = extract_entity_name(description)
        if len(entities) > 0:
            entity = max(set(entities), key = entities.count)
        else:
            entity = "Not Found"
        transaction_df.at[index,'Entity'] = entity

    # Take input the column values of cleaned entity names as a list
    cleaned_entity_names = set(transaction_df['Cleaned Entity Name'].values.tolist())
    grouped_entities = entity_name_grouper(list(cleaned_entity_names))
    # Find the group to which the entity belongs
    for index, row in transaction_df.iterrows():
        entity = row['Cleaned Entity Name']
        for key, value in grouped_entities.items():
            if entity in value:
                transaction_df.at[index, 'Reconciled Group'] = key
                #print(value)
                reconciled_name = max(value, key=len)
                transaction_df.at[index, 'Reconciled Name'] = reconciled_name
                break

    grouped_df = transaction_df.groupby('Reconciled Group')
    for name, group in grouped_df:
        # if the group is not Not Found
        if group['Reconciled Name'].iloc[0] != "Not Found" and group['Related Party'].iloc[0] == 'not related party':
            if group.loc[group['Direction'] == 'outward'].shape[0] > 0.7*group.shape[0]:
                flag = 0
                if group.loc[group['Purpose'] == 'others'].shape[0] > 0.7*group.shape[0]:
                    if group.shape[0] > 3:
                            if group.loc[group['Direction'] == 'outward', 'Debit'].std() < 0.1*group.loc[group['Direction'] == 'outward', 'Debit'].mean():
                                group['Date'] = pd.to_datetime(group['Date'], format='%d-%m-%Y')
                                if (group['Date'].diff().dt.days > 25).all() and (group['Date'].diff().dt.days < 35).all():
                                    transaction_df.loc[group.index, 'Purpose'] = "salary"
                                    flag = 1
                if flag == 0:
                    transaction_df.loc[group.index, 'Purpose'] = "supplier"
            elif group.loc[group['Direction'] == 'inward'].shape[0] > 0.7*group.shape[0]:
                transaction_df.loc[group.index, 'Purpose'] = "customer"


    return transaction_df

# Functio to convert the column number to excel column name
def get_excel_column_name(index):
    """Convert a column index (0-based) to Excel-style column name."""
    name = ""
    while index >= 0:
        name = chr(index % 26 + ord('A')) + name
        index = index // 26 - 1
    return name


# Function to generate the overview of the transactions
def overview_generator(transaction_df):
    data = transaction_df.copy()

    grouped = data.groupby(['Year', 'Month'])

    def calculate_metrics(group):
        # Initialize dictionary to hold calculated metrics
        metrics = {}

        # average transaction value is the average of the absolute value of the transaction amount of both credit and debit
        metrics['Month Year'] = group['Month Year'].iloc[0]

        metrics['Total Tranx (Nos.)'] = group.shape[0]
        metrics['Average Credit Tranx'] = group.loc[group['Credit'] > 0, 'Credit'].mean()
        metrics['Total Credit (Nos.)'] = group.loc[group['Credit'] > 0].shape[0]
        metrics['Average Debit Tranx'] = group.loc[group['Debit'] > 0, 'Debit'].mean()
        metrics['Total Debit (Nos.)'] = group.loc[group['Debit'] > 0].shape[0]

        # Calculate total credits and debits
        total_credits = group['Credit'].sum()
        total_debits = group['Debit'].sum()

        # Calculate outward and inward cheque returns
        outward_cheque_return = group.loc[group['Cheque Return'] == 'yes', 'Debit'].sum()
        inward_cheque_return = group.loc[group['Cheque Return'] == 'yes', 'Credit'].sum()


        # Calculate gross credits and debits
        gross_credits = total_credits - outward_cheque_return
        net_debits = total_debits - inward_cheque_return

        # Calculate net credits and debits
        # loan recieved is transaction type loan  and direction inward

        loan_received = group.loc[(group['Transaction Type'] == 'loan') & (group['Direction'] == 'inward'), 'Credit'].sum()
        net_credits = gross_credits  - loan_received

        # Calculate inhouse transactions
        inhouse_credit = group.loc[group['Related Party'] == "related party", 'Credit'].sum()
        inhouse_debit = group.loc[group['Related Party'] == "related party", 'Debit'].sum()

        # Calculate net cash inflow and outflow
        net_cash_inflow = net_credits - inhouse_credit
        net_cash_outflow = net_debits - inhouse_debit

        # Fill in the metrics dictionary
        metrics['Total Credits'] = total_credits
        metrics['Outward Cheque Return'] = outward_cheque_return
        metrics['Gross Credits'] = gross_credits
        metrics['Loan Received'] = loan_received
        metrics['Net Credits'] = net_credits
        metrics['Inhouse Credit'] = inhouse_credit
        metrics['Net Cash Inflow'] = net_cash_inflow

        metrics['Total Debits'] = total_debits
        metrics['Inward Cheque Return'] = inward_cheque_return
        metrics['Net Debits'] = net_debits
        metrics['Inhouse Debit'] = inhouse_debit
        metrics['Net Cash Outflow'] = net_cash_outflow

        # Write Inward Cheque Return Nos as where it is credit and cheque return is yes
        metrics['Inward Cheque Return (Nos.)'] = np.sum((group.loc[group['Cheque Return'] == 'yes', 'Credit']>0).values)
        total_cheques_received = np.sum((group.loc[group["Transaction Type"]=='cheque','Credit']>0).values)

        # ensure that the denominator is not zero
        if total_cheques_received == 0:
            metrics["Inward Cheque Return/Total Cheques Received (%)"] = 0
        else:
            metrics["Inward Cheque Return/Total Cheques Received (%)"] = metrics['Inward Cheque Return (Nos.)']/total_cheques_received

        metrics['Outward Cheque Return (Nos.)'] = np.sum((group.loc[group['Cheque Return'] == 'yes', 'Debit']>0).values)
        total_cheques_issued = np.sum((group.loc[group["Transaction Type"]=='cheque','Debit']>0).values)

        if total_cheques_issued == 0:
            metrics["Outward Cheque Return/Total Cheques Issued (%)"] = 0
        else:
            metrics["Outward Cheque Return/Total Cheques Issued (%)"] = metrics['Outward Cheque Return (Nos.)']/total_cheques_issued


        metrics["Inward Online Return (Nos.)"] = np.sum((group.loc[group['Online Return'] == 'yes', 'Credit']>0).values)
        total_online_received = np.sum((group.loc[group["Transaction Type"].isin(online_methods),'Credit']>0).values)

        if total_online_received == 0:
            metrics["Inward Online Return/Total Online Received (%)"] = 0
        else:
            metrics["Inward Online Return/Total Online Received (%)"] = metrics["Inward Online Return (Nos.)"]/total_online_received


        metrics["Outward Online Return (Nos.)"] = np.sum((group.loc[group['Online Return'] == 'yes', 'Debit']>0).values)
        total_online_issued = np.sum((group.loc[group["Transaction Type"].isin(online_methods),'Debit']>0).values)

        if total_online_issued == 0:
            metrics["Outward Online Return/Total Online Issued (%)"] = 0
        else:
            metrics["Outward Online Return/Total Online Issued (%)"] = metrics["Outward Online Return (Nos.)"]/total_online_issued

        metrics["Inward ECS Return (Nos.)"] = np.sum((group.loc[group['ECS Return'] == 'yes', 'Credit']>0).values)
        total_ecs_received = np.sum((group.loc[group["Transaction Type"]=='ecs','Credit']>0).values)

        if total_ecs_received == 0:
            metrics["Inward ECS Return/Total ECS Received (%)"] = 0
        else:
            metrics["Inward ECS Return/Total ECS Received (%)"] = metrics["Inward ECS Return (Nos.)"]/total_ecs_received

        metrics["Inhouse Credit (Nos.)"] = np.sum((group.loc[group['Related Party'] == "related party", 'Credit']>0).values)
        if metrics['Total Credit (Nos.)'] == 0:
            metrics["Inhouse Credit/Total Credits (%)"] = 0
        else :
            metrics["Inhouse Credit/Total Credits (%)"] = metrics["Inhouse Credit (Nos.)"]/metrics['Total Credit (Nos.)']

        metrics["Inhouse Debit (Nos.)"] = np.sum((group.loc[group['Related Party'] == "related party", 'Debit']>0).values)

        if metrics['Total Debit (Nos.)'] == 0:
            metrics["Inhouse Debit/Total Debits (%)"] = 0
        else:
            metrics["Inhouse Debit/Total Debits (%)"] = metrics["Inhouse Debit (Nos.)"]/metrics['Total Debit (Nos.)']

        metrics["Loan Repaid"] = group.loc[(group['Transaction Type'] == 'loan') & (group['Direction'] == 'outward'), 'Debit'].sum()
        metrics["ECS Payment"] = group.loc[group['Transaction Type'] == 'ecs', 'Debit'].sum()

        metrics["No. of Unique ECS/EMI's"] = np.sum((group['Transaction Type'] == 'ecs').values)

        metrics["Interest Paid"] = group.loc[group['Purpose'] == 'interest', 'Debit'].sum()
        # Return the metrics dictionary as a Series
        return pd.Series(metrics)

    # Apply the function to each group (Year-Month) and create a new DataFrame
    result_df = grouped.apply(calculate_metrics).reset_index()

    # Remove the month and year columns
    result_df = result_df.drop(columns=['Year', 'Month'])
    # replace nan with 0

    result_df = result_df.fillna(0)

    #Make month year as index
    result_df = result_df.set_index('Month Year')

    # Add a column with index name Overall/Total at the end which takes sum  of all the row values where the column is not average and take average where the column is average

    if  result_df.shape[0] >= 1:
        result_df.loc['Overall/Total'] = result_df.sum()

        # For the columns which are average, calculate the average of the values in the column
        result_df.loc['Overall/Total','Average Credit Tranx'] = result_df.loc['Overall/Total','Average Credit Tranx']/(result_df.shape[0]-1)
        result_df.loc['Overall/Total','Average Debit Tranx'] = result_df.loc['Overall/Total','Average Debit Tranx']/(result_df.shape[0]-1)

        # For the columns which are percentage, calculate the average of the values in the column
        result_df.loc['Overall/Total',"Inward Cheque Return/Total Cheques Received (%)"] = result_df.loc['Overall/Total',"Inward Cheque Return/Total Cheques Received (%)"]/(result_df.shape[0]-1)
        result_df.loc['Overall/Total',"Outward Cheque Return/Total Cheques Issued (%)"] = result_df.loc['Overall/Total',"Outward Cheque Return/Total Cheques Issued (%)"]/(result_df.shape[0]-1)
        result_df.loc['Overall/Total',"Inward Online Return/Total Online Received (%)"] = result_df.loc['Overall/Total',"Inward Online Return/Total Online Received (%)"]/(result_df.shape[0]-1)
        result_df.loc['Overall/Total',"Outward Online Return/Total Online Issued (%)"] = result_df.loc['Overall/Total',"Outward Online Return/Total Online Issued (%)"]/(result_df.shape[0]-1)
        result_df.loc['Overall/Total',"Inward ECS Return/Total ECS Received (%)"] = result_df.loc['Overall/Total',"Inward ECS Return/Total ECS Received (%)"]/(result_df.shape[0]-1)
        result_df.loc['Overall/Total',"Inhouse Credit/Total Credits (%)"] = result_df.loc['Overall/Total',"Inhouse Credit/Total Credits (%)"]/(result_df.shape[0]-1)
        result_df.loc['Overall/Total',"Inhouse Debit/Total Debits (%)"] = result_df.loc['Overall/Total',"Inhouse Debit/Total Debits (%)"]/(result_df.shape[0]-1)


    # Rearrange this to the top row -
    result_df = result_df.reindex(index = ['Overall/Total'] + list(result_df.index[:-1]))

    # Take transpose of the dataframe with indexs as columns
    result_df = result_df.T
    return result_df

def statement_details_formatter(statement_details,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)

    # Create a new worksheet
    worksheet = workbook.create_sheet(title='Statement Details')
    # Make it the active sheet
    workbook.active = worksheet

    start_row,start_col = 3,3

    rows = dataframe_to_rows(statement_details, header=True)

    # Do the formatting for the first row same as the other rows
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True,color='FFFFFF')
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
            else:
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=False)

            # Make all thin borders
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Dont show gridlines
    worksheet.sheet_view.showGridLines = False
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length + 5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    workbook.active = 0
    workbook.save(folder_path+main_report_name)

# Function to format and save the overview sheet
def overview_sheet_formatter(overview_sheet,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    # Create a new worksheet
    worksheet = workbook.create_sheet(title='Overview')
    # Make it the active sheet
    workbook.active = worksheet

    start_row,start_col = 3,3

    rows = dataframe_to_rows(overview_sheet, index=True, header=True)

    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True)
            else:
                if c_idx == start_col:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left')
    worksheet.delete_rows(4)
    # Define a list of row indices after which empty rows should be inserted
    positions = [4,9,16,21,31,35]

    font_1 = Font(bold=True, underline='single', color='FFFFFF', size=11)

    # Iterate over the positions in reverse order to avoid affecting row indices
    for pos in sorted(positions, reverse=True):
        # Insert an empty row at the specified position
        worksheet.insert_rows(pos)
        for col in range(overview_sheet.shape[1]+1):
            cell = worksheet.cell(row=pos, column=col + 3)
            cell.font = font_1
            cell.fill = PatternFill(start_color='4C545F', end_color='4C545F', fill_type='solid')


    positions = [4,10,18,24,35,40]
    for row in worksheet.iter_rows(min_row=3,max_row=44, min_col=3, max_col=overview_sheet.shape[1] + 3):
        for cell in row:
            cell.border = Border()

    for pos in sorted([3]+positions, reverse=True):
        for col in range(overview_sheet.shape[1]+1):
            cell = worksheet.cell(row=pos, column=col + 3)
            cell.border = Border(top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet['C2'] = 'Overview of Bank Statement'
    worksheet['C3'] = 'Particulars (INR)'
    worksheet['C3'].font = Font(bold=True)

    # Create a format for the merged cell
    font = Font(bold=True, underline='single', color='FFFFFF', size=14)
    fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    alignment = Alignment(horizontal='center')

    # Apply formatting to the merged cell
    cell = worksheet['C2']
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    # thick borders on the whole merged cell
    cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Make the cells C5 to C44 left aligned
    for row in worksheet.iter_rows(min_row=5, max_row=44, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # Make the cells D5 to D44 bold
    for row_num,row in enumerate(worksheet.iter_rows(min_row=5, max_row=44, min_col=4, max_col=4)):
        for cell in row:
            if row_num+5 not in [4,10,18,24,35,40]:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            if row_num+5 in [6,8]:
                cell.font = Font(bold=True)
            else :
                cell.font = Font(bold=True,italic=True)

    # For row 3 make bg color #6CC9E5
    for row in worksheet.iter_rows(min_row=3, max_row=3, min_col=3, max_col=overview_sheet.shape[1] + 3):
        for cell in row:
            cell.fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')

    red_colored_rows = [7,9,12,14,16,20,22,25,27,29,31,33,36,38,41,42,43]

    for row_num in red_colored_rows:
        for row in worksheet.iter_rows(min_row=row_num, max_row=row_num, min_col=3, max_col=overview_sheet.shape[1] + 3):
            for cell in row:
                cell.font = Font(color='EF2E00')

    for col in worksheet.iter_cols(min_row=3, max_row=44, min_col=4, max_col=overview_sheet.shape[1] + 3):
        for cell in col:
            if cell.row in [4,10,18,24,35,40]:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
            else:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

    important_rows = [11,13,15,17,19,21,23]
    # shade the important row cells with grey
    for row_num in important_rows:
        for row in worksheet.iter_rows(min_row=row_num, max_row=row_num, min_col=3, max_col=overview_sheet.shape[1] + 3):
            for cell in row:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    merged_cell = worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=overview_sheet.shape[1] + 3)

    # make a thick border covering the whole table
    for row in worksheet.iter_rows(min_row=3, max_row=44, min_col=3, max_col=overview_sheet.shape[1] + 3):
        #get the first and last cell of the row
        first_cell = row[0]
        last_cell = row[-1]

        if first_cell.row in [3]+positions:
            first_cell.border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
            last_cell.border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        else:
            first_cell.border = Border(left=Side(style='thick'), right=Side(style='thin'))
            last_cell.border = Border(left=Side(style='thin'), right=Side(style='thick'))

        if row[0].row == 44:
            for cell in row:
                if cell.col_idx == 3:
                    cell.border = Border(left=Side(style='thick'), right=Side(style='thin'),bottom=Side(style='thick'))
                elif cell.col_idx == overview_sheet.shape[1] + 3:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thick'),bottom=Side(style='thick'))
                else :
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),bottom=Side(style='thick'))

    # Dont show gridlines
    worksheet.sheet_view.showGridLines = False
    # Save the modified workbook
    # Replace value which are 0 with -
    percentage_rows = [26,28,30,32,34,37,39]
    numeric_rows = [5,7,9,25,27,29,31,33,36,38,43]
    amount_rows = [6,8,10,11,12,13,14,15,16,17,18,19,20,21,22,23,41,42,44]
    for row in worksheet.iter_rows(min_row=5, max_row=44, min_col=4, max_col=overview_sheet.shape[1] + 3):
        for cell in row:
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='right')
            else:
                if cell.row in percentage_rows:
                    cell.number_format = '0.00%'

                elif cell.row in numeric_rows:
                    cell.number_format = '#,##0'
                elif cell.row in amount_rows:
                    cell.number_format = '#,##0.00'

    # do auto fit for all columns
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length + 5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    workbook.active = 0
    workbook.save(folder_path+main_report_name)

# Function to generate the summary of debits and credits
def summary_of_debits_and_credits(transaction_df):
    data = transaction_df.copy()

    grouped = data.groupby(['Year', 'Month'])

    def calculate_metrics(group):
        # Initialize dictionary to hold calculated metrics
        metrics = {}

        metrics['Month Year'] = group['Month Year'].iloc[0]

        # here the values should not consider related party
        metrics["Cash Deposit"] = group.loc[(group['Transaction Type'] == 'cash') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["Cheque Receipt"] = group.loc[(group['Transaction Type'] == 'cheque') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["Online Receipt"] = group.loc[(group['Transaction Type'].isin(online_methods)) & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()

        # all receipts other than cash,cheque and online
        metrics["Other Receipt"] = group.loc[(~group['Transaction Type'].isin(["cash","cheque"] + online_methods)) & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()

        # write for all the receipts for remaining transaction types
        metrics["Internet Banking Receipt"] = group.loc[(group['Transaction Type'] == 'internet banking') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["RTGS Receipt"] = group.loc[(group['Transaction Type'] == 'rtgs') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["NEFT Receipt"] = group.loc[(group['Transaction Type'] == 'neft') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["UPI Receipt"] = group.loc[(group['Transaction Type'] == 'upi') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["IMPS Receipt"] = group.loc[(group['Transaction Type'] == 'imps') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["ECS Receipt"] = group.loc[(group['Transaction Type'] == 'ecs') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()
        metrics["Transfer Receipt"] = group.loc[(group['Transaction Type'] == 'transfer') & (group['Direction'] == 'inward') & (group['Related Party'] == "not related party"), 'Credit'].sum()

        metrics["Inhouse Receipt"] = group.loc[(group['Related Party'] == "related party") & (group['Direction'] == 'inward'), 'Credit'].sum()

        metrics["Total Receipt"] = group.loc[(group['Direction'] == 'inward'), 'Credit'].sum()

        metrics["No. of Cash Deposits"] = np.sum((group.loc[(group['Transaction Type'] == 'cash') & (group['Related Party'] == "not related party"), 'Credit']>0).values)
        metrics["No. of Cheque Receipts"] = np.sum((group.loc[(group['Transaction Type'] == 'cheque') & (group['Related Party'] == "not related party"), 'Credit']>0).values)
        metrics["No. of Online Receipts"] = np.sum((group.loc[(group['Transaction Type'].isin(online_methods)) & (group['Related Party'] == "not related party"), 'Credit']>0).values)
        metrics["No. of Other Receipts"] = np.sum((group.loc[(~group['Transaction Type'].isin(["cash","cheque"] + online_methods)) & (group['Related Party'] == "not related party"),'Credit']>0).values)
        metrics["No. of Inhouse Receipts"] = np.sum((group.loc[group['Related Party'] == "related party", 'Credit']>0).values)
        metrics["No. of Total Receipts"] = np.sum((group['Direction'] == 'inward').values)

        metrics["Cash Withdrawals"] = group.loc[(group['Transaction Type'] == 'cash') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["Cheque Payment"] = group.loc[(group['Transaction Type'] == 'cheque') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["Online Payment"] = group.loc[(group['Transaction Type'].isin(online_methods)) & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()

        # all payments other than cash,cheque and online
        metrics["Other Payment"] = group.loc[(~group['Transaction Type'].isin(["cash","cheque"] + online_methods)) & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()

        # write for all the payments for remaining transaction types
        metrics["Internet Banking Payment"] = group.loc[(group['Transaction Type'] == 'internet banking') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["RTGS Payment"] = group.loc[(group['Transaction Type'] == 'rtgs') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["NEFT Payment"] = group.loc[(group['Transaction Type'] == 'neft') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["UPI Payment"] = group.loc[(group['Transaction Type'] == 'upi') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["IMPS Payment"] = group.loc[(group['Transaction Type'] == 'imps') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["ECS Payment"] = group.loc[(group['Transaction Type'] == 'ecs') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()
        metrics["Transfer Payment"] = group.loc[(group['Transaction Type'] == 'transfer') & (group['Direction'] == 'outward') & (group['Related Party'] == "not related party"), 'Debit'].sum()

        metrics["Inhouse Payment"] = group.loc[(group['Related Party'] == "related party") & (group['Direction'] == 'outward'), 'Debit'].sum()

        metrics["Total Payment"] = group.loc[(group['Direction'] == 'outward'), 'Debit'].sum()

        metrics["No. of Cash Withdrawals"] = np.sum((group.loc[(group['Transaction Type'] == 'cash') & (group['Related Party'] == "not related party"), 'Debit']>0).values)

        metrics["No. of Cheque Payments"] = np.sum((group.loc[(group['Transaction Type'] == 'cheque') & (group['Related Party'] == "not related party"), 'Debit']>0).values)

        metrics["No. of Online Payments"] = np.sum((group.loc[(group['Transaction Type'].isin(online_methods)) & (group['Related Party'] == "not related party"), 'Debit']>0).values)

        metrics["No. of Other Payments"] = np.sum((group.loc[(~group['Transaction Type'].isin(["cash","cheque"] + online_methods)) & (group['Related Party'] == "not related party"), 'Debit']>0).values)

        metrics["No. of Inhouse Payments"] = np.sum((group.loc[group['Related Party'] == "related party", 'Debit']>0).values)

        metrics["No. of Total Payments"] = np.sum((group['Direction'] == 'outward').values)

        return pd.Series(metrics)

    # Apply the function to each group (Year-Month) and create a new DataFrame
    result_df = grouped.apply(calculate_metrics).reset_index()

    # Remove the month and year columns
    result_df = result_df.drop(columns=['Year', 'Month'])
    # replace nan with 0

    result_df = result_df.fillna(0)

    #Make month year as index
    result_df = result_df.set_index('Month Year')

    if  result_df.shape[0] >= 1:
        result_df.loc['Overall/Total'] = result_df.sum()

    # Rearrange this to the top row -
    result_df = result_df.reindex(index = ['Overall/Total'] + list(result_df.index[:-1]))

    # Take transpose of the dataframe with indexs as columns
    result_df = result_df.T
    return result_df

# Function to format and save the summary of debits and credits sheet
def summary_sheet_formatter(summary_sheet,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Summary of Debits and Credits')
    workbook.active = worksheet

    start_row,start_col = 3,3

    rows = dataframe_to_rows(summary_sheet, index=True, header=True)

    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True)
            else:
                if c_idx == start_col:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left')
    worksheet.delete_rows(4)

    # Define a list of row indices after which empty rows should be inserted
    positions = [4,17,23,36]

    font_1 = Font(bold=True, underline='single', color='FFFFFF', size=11)

    # Iterate over the positions in reverse order to avoid affecting row indices
    for pos in sorted(positions, reverse=True):
        # Insert an empty row at the specified position
        worksheet.insert_rows(pos)
        for col in range(summary_sheet.shape[1]+1):
            cell = worksheet.cell(row=pos, column=col + 3)
            cell.font = font_1
            cell.fill = PatternFill(start_color='4C545F', end_color='4C545F', fill_type='solid')

    positions = [4,18,25,39]
    end_row = 45
    for row in worksheet.iter_rows(min_row=3,max_row=end_row, min_col=3, max_col=summary_sheet.shape[1] + 3):
        for cell in row:
            cell.border = Border()

    for pos in sorted([3]+positions, reverse=True):
        for col in range(summary_sheet.shape[1]+1):
            cell = worksheet.cell(row=pos, column=col + 3)
            cell.border = Border(top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet['C2'] = 'Summary of Debits and Credits'
    worksheet['C3'] = 'Months'
    worksheet['C3'].font = Font(bold=True)

    # Create a format for the merged cell
    font = Font(bold=True, underline='single', color='FFFFFF', size=14)
    fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    alignment = Alignment(horizontal='center')

    # Apply formatting to the merged cell
    cell = worksheet['C2']
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    # thick borders on the whole merged cell
    cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))


    # Make the cells C5 to C44 left aligned
    for row in worksheet.iter_rows(min_row=5, max_row=end_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # Make the cells D5 to D31 bold
    for row_num,row in enumerate(worksheet.iter_rows(min_row=5, max_row=end_row, min_col=4, max_col=4)):
        for cell in row:
            if row_num+5 not in positions:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


    # For row 3 make bg color #6CC9E5
    for row in worksheet.iter_rows(min_row=3, max_row=3, min_col=3, max_col=summary_sheet.shape[1] + 3):
        for cell in row:
            cell.fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')

    red_colored_rows = [16,23,37,44]

    for row_num in red_colored_rows:
        for row in worksheet.iter_rows(min_row=row_num, max_row=row_num, min_col=4, max_col=summary_sheet.shape[1] + 3):
            for cell in row:
                cell.font = Font(color='0967BA',italic=True)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    for col in worksheet.iter_cols(min_row=3, max_row=end_row, min_col=4, max_col=summary_sheet.shape[1] + 3):
        for cell in col:
            if cell.row in positions:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
            else:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

    merged_cell = worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=summary_sheet.shape[1] + 3)

    # make a thick border covering the whole table
    for row in worksheet.iter_rows(min_row=3, max_row=end_row, min_col=3, max_col=summary_sheet.shape[1] + 3):
        #get the first and last cell of the row
        first_cell = row[0]
        last_cell = row[-1]

        if first_cell.row in [3]+positions:
            first_cell.border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
            last_cell.border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        else:
            first_cell.border = Border(left=Side(style='thick'), right=Side(style='thin'))
            last_cell.border = Border(left=Side(style='thin'), right=Side(style='thick'))

        if row[0].row == end_row:
            for cell in row:
                if cell.col_idx == 3:
                    cell.border = Border(left=Side(style='thick'), right=Side(style='thin'),bottom=Side(style='thick'))
                elif cell.col_idx == summary_sheet.shape[1] + 3:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thick'),bottom=Side(style='thick'))
                else :
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),bottom=Side(style='thick'))

    # Dont show gridlines
    worksheet.sheet_view.showGridLines = False
    # Save the modified workbook
    # Replace value which are 0 with -
    numeric_rows = [19,20,21,22,23,24,40,41,42,43,44,45]
    amount_rows = [5,6,7,8,9,10,11,12,13,14,15,16,17,26,27,28,29,30,31,32,33,34,35,36,37,38]
    for row in worksheet.iter_rows(min_row=5, max_row=end_row, min_col=4, max_col=summary_sheet.shape[1] + 3):
        for cell in row:
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='right')
            else:
                if cell.row in numeric_rows:
                    cell.number_format = '#,##0'
                elif cell.row in amount_rows:
                    cell.number_format = '#,##0.00'

    worksheet['C4'] = "INFLOWS"
    worksheet['C4'].font = Font(bold=True,size=13,color='FFFFFF')
    worksheet['C4'].alignment = Alignment(horizontal='center')

    worksheet['C25'] = "OUTFLOWS"
    worksheet['C25'].font = Font(bold=True,size=13,color='FFFFFF')
    worksheet['C25'].alignment = Alignment(horizontal='center')

    worksheet['D4'] = "Value"
    worksheet['D4'].font = Font(bold=True,color='FFFFFF')
    worksheet['D4'].alignment = Alignment(horizontal='center')

    worksheet['D25'] = "Value"
    worksheet['D25'].font = Font(bold=True,color='FFFFFF')
    worksheet['D25'].alignment = Alignment(horizontal='center')

    worksheet['D18'] = "(No.)"
    worksheet['D18'].font = Font(bold=True,color='FFFFFF')
    worksheet['D18'].alignment = Alignment(horizontal='center')

    worksheet['D39'] = "(No.)"
    worksheet['D39'].font = Font(bold=True,color='FFFFFF')
    worksheet['D39'].alignment = Alignment(horizontal='center')

    # do auto fit for all columns
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length + 5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    worksheet.row_dimensions.group(5,16,outline_level=1)
    worksheet.row_dimensions.group(26,37,outline_level=1)

    worksheet.row_dimensions.group(9,15,outline_level=2,hidden=True)
    worksheet.row_dimensions.group(30,36,outline_level=2,hidden=True)

    workbook.active = 0
    worksheet.title = 'Summary of Credits and Debits'
    workbook.save(folder_path+main_report_name)

# Function to generate the top credits and debits
def top_credits_and_debits(transaction_df):
    data = transaction_df.copy()

    # Make 2 dataframes seperately for top 10 credits and debits in descending order of the amount

    # Top 10 Credits if there are atleast 10 credits else take all credits
    if data.loc[data['Direction'] == 'inward'].shape[0] >= 10:
        top_credits = data.loc[data['Direction'] == 'inward', ['Date','Description','Purpose', 'Credit']].sort_values(by='Credit', ascending=False).head(10)
    else:
        top_credits = data.loc[data['Direction'] == 'inward', ['Date','Description','Purpose', 'Credit']].sort_values(by='Credit', ascending=False)
        top_credits.reset_index(drop=True, inplace=True)
        for i in range(10 - top_credits.shape[0]):
            top_credits.loc[top_credits.shape[0]] = [np.nan]*top_credits.shape[1]


    # Top 10 Debits
    if data.loc[data['Direction'] == 'outward'].shape[0] >= 10:
        top_debits = data.loc[data['Direction'] == 'outward', ['Date','Description','Purpose', 'Debit']].sort_values(by='Debit', ascending=False).head(10)
    else:
        top_debits = data.loc[data['Direction'] == 'outward', ['Date','Description','Purpose', 'Debit']].sort_values(by='Debit', ascending=False)
        top_debits.reset_index(drop=True, inplace=True)
        for i in range(10 - top_debits.shape[0]):
            top_debits.loc[top_debits.shape[0]] = [np.nan]*top_debits.shape[1]


    # Now doing the same monthwise
    grouped = data.groupby(['Year', 'Month'])

    def calculate_top_credits(group):

        # Initialize dictionary to hold calculated metrics
        metrics = {}

        metrics['Month Year'] = group['Month Year'].iloc[0]

        # Top 10 Credits

        top_credits = group.loc[group['Direction'] == 'inward', ['Date','Description','Purpose', 'Credit']].sort_values(by='Credit', ascending=False)
        # reindex
        top_credits.reset_index(drop=True, inplace=True)
        if top_credits.shape[0] >= 5:
            top_credits = top_credits.head(5)
        else :
            # add empty rows to make it 5
            for i in range(5 - top_credits.shape[0]):
                top_credits.loc[top_credits.shape[0]] = [np.nan]*top_credits.shape[1]

        # Add a row with the first col value as Month Year in top_credits
        top_credits.loc[-1] = [metrics['Month Year']] + [np.nan]*3
        top_credits.index = top_credits.index + 1
        top_credits = top_credits.sort_index()

        # Top 10 Debits
        top_debits = group.loc[group['Direction'] == 'outward', ['Date','Description','Purpose', 'Debit']].sort_values(by='Debit', ascending=False)
        # reindex
        top_debits.reset_index(drop=True, inplace=True)
        if top_debits.shape[0] >= 5:
            top_debits = top_debits.head(5)
        else:
            # add empty rows to make it 5
            for i in range(5 - top_debits.shape[0]):
                top_debits.loc[top_debits.shape[0]] = [np.nan]*top_debits.shape[1]
        # Add a row with the first col value as Month Year in top_debits
        top_debits.loc[-1] = [metrics['Month Year']] + [np.nan]*3
        top_debits.index = top_debits.index + 1
        top_debits = top_debits.sort_index()


        # Return the metrics dictionary as a Series
        return pd.Series({'Top 10 Credits': top_credits, 'Top 10 Debits': top_debits})

    # Apply the function to each group (Year-Month) and create a new DataFrame
    result_df = grouped.apply(calculate_top_credits).reset_index()

    # Remove the month and year columns
    result_df = result_df.drop(columns=['Year', 'Month'])

    return top_credits,top_debits,result_df

# Function to generate the loan transactions
def loan_transactions(transaction_df):
    data = transaction_df.copy()
    grouped = data.groupby(['Year', 'Month'])

    def calculate_loan_transactions(group):
        # Initialize dictionary to hold calculated metrics
        metrics = {}

        metrics['Month Year'] = group['Month Year'].iloc[0]

        # Identify those transactions whose Purpose value is loan

        loan_transactions = group.loc[group['Purpose']=='loan', ['Date','Description','Credit', 'Debit']]
        # reindex
        loan_transactions.reset_index(drop=True, inplace=True)

        # Add row with the first column value as Month Year
        loan_transactions.loc[-1] = [metrics['Month Year']] + [np.nan]*3
        loan_transactions.index = loan_transactions.index + 1
        loan_transactions = loan_transactions.sort_index()
        # return the loan transactions
        return pd.Series({'Loan Transactions': loan_transactions})

    # Apply the function to each group (Year-Month) and create a new DataFrame
    loan_transction_df = grouped.apply(calculate_loan_transactions).reset_index()

    # Remove the month and year columns
    loan_transction_df = loan_transction_df.drop(columns=['Year', 'Month'])

    return loan_transction_df

# Function to generate the interest transactions
def interest_transactions(transaction_df):
    data = transaction_df.copy()
    grouped = data.groupby(['Year', 'Month'])

    def calculate_interest_transactions(group):
        # Initialize dictionary to hold calculated metrics
        metrics = {}

        metrics['Month Year'] = group['Month Year'].iloc[0]

        # Identify those transactions whose Purpose value is loan

        interest_transactions = group.loc[group['Purpose']=='interest', ['Date','Description','Credit', 'Debit']]
        # reindex
        interest_transactions.reset_index(drop=True, inplace=True)

        # Add row with the first column value as Month Year
        interest_transactions.loc[-1] = [metrics['Month Year']] + [np.nan]*3
        interest_transactions.index = interest_transactions.index + 1
        interest_transactions = interest_transactions.sort_index()

        # return the loan transactions
        return pd.Series({'Interest Transactions': interest_transactions})

    # Apply the function to each group (Year-Month) and create a new DataFrame
    interest_transction_df = grouped.apply(calculate_interest_transactions).reset_index()

    # Remove the month and year columns
    interest_transction_df = interest_transction_df.drop(columns=['Year', 'Month'])

    return interest_transction_df

# Function to format and save the top credits and debits sheet
def top_credit_debit_formatter(top_credits,top_debits,top_credits_debits_monthwise,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Top Credits and Debits')
    workbook.active = worksheet
    # create a new sheet
    worksheet = workbook.active

    rows = dataframe_to_rows(top_credits, index=False, header=True)
    top_credits_height, top_credits_width = top_credits.shape

    start_row,start_col = 2,2
    # Merge cells from B2 till the width of top_credits and write "Top 10 Credit Transactions"
    worksheet.cell(row=start_row, column=start_col, value="Top 10 Credit Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+top_credits_width-1)


    # write the top_credits datafram at the cell B3
    for r_idx, row in enumerate(rows, start=start_row+1):
        if r_idx == start_row+1:
            for c_idx, value in enumerate(row, start=start_col):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).font = Font(bold=True,color='FFFFFF')
                worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                worksheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

        for c_idx, value in enumerate(row, start=start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == start_col + top_credits_width - 1:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '0.00'

            if r_idx == start_row + top_credits_height+1:
                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))

            else:
                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), top=Side(style='thin'), right=Side(style='thin'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'))

    rows = dataframe_to_rows(top_debits, index=False, header=True)
    top_credits_height, top_credits_width = top_credits.shape

    start_row,start_col = 2,8
    # Merge cells from B2 till the width of top_credits and write "Top 10 Credit Transactions"
    worksheet.cell(row=start_row, column=start_col, value="Top 10 Debit Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+top_credits_width-1)


    # write the top_credits datafram at the cell B3
    for r_idx, row in enumerate(rows, start=start_row+1):
        if r_idx == start_row+1:
            for c_idx, value in enumerate(row, start=start_col):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).font = Font(bold=True,color='FFFFFF')
                worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                worksheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

        for c_idx, value in enumerate(row, start=start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == start_col + top_credits_width - 1:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '0.00'

            if r_idx == start_row + top_credits_height+1:
                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))

            else:
                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), top=Side(style='thin'), right=Side(style='thin'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'))

    start_row_for_monthly,start_col = top_credits_height+6,2

    top_credits_monthly_list = top_credits_debits_monthwise["Top 10 Credits"]

    worksheet.cell(row=start_row_for_monthly, column=start_col, value="Top 5 Credits Monthwise")
    worksheet.cell(row=start_row_for_monthly, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row_for_monthly, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row_for_monthly, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row_for_monthly, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row_for_monthly, start_column=start_col, end_row=start_row_for_monthly, end_column=start_col+top_credits_width-1)

    current_row = start_row_for_monthly+1
    month_name_list = []
    curr_year,curr_year_idx,month_start_list = None,None,[]
    for i in range(len(top_credits_monthly_list)):
        # Write the column names only for the first time
        if i == 0:
            # Add a row by taking the column names ofmfirst dataframe
            col_name_list = top_credits_monthly_list[i].columns.tolist()
            for c_idx, value in enumerate(col_name_list, start=start_col):
                worksheet.cell(row=current_row, column=c_idx, value=value)
                worksheet.cell(row=current_row, column=c_idx).font = Font(bold=True,color='FFFFFF')
                worksheet.cell(row=current_row, column=c_idx).alignment = Alignment(horizontal='center')
                worksheet.cell(row=current_row, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
                if c_idx == start_col:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            current_row += 1

        rows = dataframe_to_rows(top_credits_monthly_list[i], index=False, header=False)

        top_credits_monthly_height, top_credits_monthly_width = top_credits_monthly_list[i].shape

        start_row = current_row
        nan_rows = 0

        for r_idx, row in enumerate(rows, start=start_row):
            month_start_list.append(r_idx+1)
            if r_idx == start_row:
                for c_idx, value in enumerate(row, start=start_col):
                    if c_idx == start_col :
                        if curr_year == None:
                            #convert month year to year
                            curr_year = value.split(" ")[1]
                            curr_year_idx = r_idx
                            worksheet.cell(row=r_idx,column=start_col-1,value=curr_year)
                            # fill yellow highlight
                            worksheet.cell(row=r_idx,column=start_col-1).fill= PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        else :
                            if curr_year != value.split(" ")[1]:
                                curr_year = value.split(" ")[1]
                                worksheet.cell(row=r_idx,column=start_col-1,value=curr_year)
                                worksheet.cell(row=r_idx,column=start_col-1).fill= PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                                worksheet.row_dimensions.group(curr_year_idx+1, r_idx-1, hidden=True, outline_level=1)
                                curr_year_idx = r_idx
                    month_name_list.append(r_idx)
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
                    worksheet.cell(row=r_idx, column=c_idx).font = Font(bold=True,color='FFFFFF')
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                    worksheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')

                    if c_idx == start_col:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                    elif c_idx == start_col + top_credits_monthly_width - 1:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                    else:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            for c_idx, value in enumerate(row, start=start_col):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                if c_idx == start_col + top_credits_monthly_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '0.00'

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
                elif c_idx == start_col + top_credits_monthly_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
            current_row += 1

            worksheet.row_dimensions.group(curr_year_idx+1, current_row-1, hidden=True, outline_level=1)

    # Merge all the cells of the month name rows
    for i in range(len(month_name_list)):
        worksheet.cell(row=month_name_list[i], column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=month_name_list[i], column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=month_name_list[i], start_column=start_col, end_row=month_name_list[i], end_column=start_col+top_credits_monthly_width-1)

    start_row_for_monthly,start_col = top_credits_height+6,8

    top_credits_monthly_list = top_credits_debits_monthwise["Top 10 Debits"]

    worksheet.cell(row=start_row_for_monthly, column=start_col, value="Top 5 Debits Monthwise")
    worksheet.cell(row=start_row_for_monthly, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row_for_monthly, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row_for_monthly, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row_for_monthly, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row_for_monthly, start_column=start_col, end_row=start_row_for_monthly, end_column=start_col+top_credits_width-1)

    current_row = start_row_for_monthly+1
    month_name_list = []
    for i in range(len(top_credits_monthly_list)):
        # Write the column names only for the first time
        if i == 0:
            # Add a row by taking the column names ofmfirst dataframe
            col_name_list = top_credits_monthly_list[i].columns.tolist()
            for c_idx, value in enumerate(col_name_list, start=start_col):
                worksheet.cell(row=current_row, column=c_idx, value=value)
                worksheet.cell(row=current_row, column=c_idx).font = Font(bold=True,color='FFFFFF')
                worksheet.cell(row=current_row, column=c_idx).alignment = Alignment(horizontal='center')
                worksheet.cell(row=current_row, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
                if c_idx == start_col:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                elif c_idx == start_col + top_credits_width - 1:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            current_row += 1

        rows = dataframe_to_rows(top_credits_monthly_list[i], index=False, header=False)

        top_credits_monthly_height, top_credits_monthly_width = top_credits_monthly_list[i].shape

        start_row = current_row
        nan_rows = 0

        for r_idx, row in enumerate(rows, start=start_row):
            if r_idx == start_row:
                for c_idx, value in enumerate(row, start=start_col):
                    month_name_list.append(r_idx)
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
                    worksheet.cell(row=r_idx, column=c_idx).font = Font(bold=True,color='FFFFFF')
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                    worksheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')

                    if c_idx == start_col:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                    elif c_idx == start_col + top_credits_monthly_width - 1:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                    else:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            for c_idx, value in enumerate(row, start=start_col):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                if c_idx == start_col + top_credits_monthly_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '0.00'

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
                elif c_idx == start_col + top_credits_monthly_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
            current_row += 1

    # Merge all the cells of the month name rows
    for i in range(len(month_name_list)):
        worksheet.cell(row=month_name_list[i], column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=month_name_list[i], column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=month_name_list[i], start_column=start_col, end_row=month_name_list[i], end_column=start_col+top_credits_monthly_width-1)

    # do auto fit for all columns
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+3)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    # remove gridlines

    worksheet.sheet_view.showGridLines = False
    workbook.active = 0
    workbook.save(folder_path+main_report_name)

# Function to format and save the loan transactions sheet
def loan_transactions_formatter(loan_transction_df,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Loan Transactions')
    workbook.active = worksheet
    # create a new sheet
    worksheet = workbook.active

    start_row_for_monthly,start_col = 2,2

    loan_transction_list = loan_transction_df["Loan Transactions"]
    if len(loan_transction_list) > 0:
        loan_transaction_height, loan_transaction_width = loan_transction_list[0].shape
    else :
        loan_transaction_height, loan_transaction_width = 0,4
    worksheet.cell(row=start_row_for_monthly, column=start_col, value="Loan Transactions")
    worksheet.cell(row=start_row_for_monthly, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row_for_monthly, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row_for_monthly, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row_for_monthly, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row_for_monthly, start_column=start_col, end_row=start_row_for_monthly, end_column=start_col+loan_transaction_width-1)

    current_row = start_row_for_monthly+1
    month_name_list = []
    curr_year,curr_year_idx,month_start_list = None,None,[]
    for i in range(len(loan_transction_list)):
        loan_transaction_height, loan_transaction_width = loan_transction_list[i].shape
        # Write the column names only for the first time
        if i == 0:
            # Add a row by taking the column names ofmfirst dataframe
            col_name_list = loan_transction_list[i].columns.tolist()
            for c_idx, value in enumerate(col_name_list, start=start_col):
                worksheet.cell(row=current_row, column=c_idx, value=value)
                worksheet.cell(row=current_row, column=c_idx).font = Font(bold=True,color='FFFFFF')
                worksheet.cell(row=current_row, column=c_idx).alignment = Alignment(horizontal='center')
                worksheet.cell(row=current_row, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
                if c_idx == start_col:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                elif c_idx == start_col + loan_transaction_width - 1:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            current_row += 1

        rows = dataframe_to_rows(loan_transction_list[i], index=False, header=False)

        start_row = current_row
        nan_rows = 0

        for r_idx, row in enumerate(rows, start=start_row):
            month_start_list.append(r_idx+1)
            if r_idx == start_row:
                for c_idx, value in enumerate(row, start=start_col):
                    if c_idx == start_col :
                        if curr_year == None:
                            #convert month year to year
                            curr_year = value.split(" ")[1]
                            curr_year_idx = r_idx
                            worksheet.cell(row=r_idx,column=start_col-1,value=curr_year)
                            # fill yellow highlight
                            worksheet.cell(row=r_idx,column=start_col-1).fill= PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        else :
                            if curr_year != value.split(" ")[1]:
                                curr_year = value.split(" ")[1]
                                worksheet.cell(row=r_idx,column=start_col-1,value=curr_year)
                                worksheet.cell(row=r_idx,column=start_col-1).fill= PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                                worksheet.row_dimensions.group(curr_year_idx+1, r_idx-1, hidden=True, outline_level=1)
                                curr_year_idx = r_idx
                    month_name_list.append(r_idx)
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
                    worksheet.cell(row=r_idx, column=c_idx).font = Font(bold=True,color='FFFFFF')
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                    worksheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')

                    if c_idx == start_col:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                    elif c_idx == start_col + loan_transaction_width - 1:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                    else:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            for c_idx, value in enumerate(row, start=start_col):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                if c_idx >= start_col + loan_transaction_width - 2:
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '0.00'

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
                elif c_idx == start_col + loan_transaction_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
            current_row += 1

            worksheet.row_dimensions.group(curr_year_idx+1, current_row-1, hidden=True, outline_level=1)

    # Merge all the cells of the month name rows
    for i in range(len(month_name_list)):
        worksheet.cell(row=month_name_list[i], column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=month_name_list[i], column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=month_name_list[i], start_column=start_col, end_row=month_name_list[i], end_column=start_col+loan_transaction_width-1)

    # do auto fit for all columns
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+3)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    # remove gridlines

    worksheet.sheet_view.showGridLines = False
    workbook.active = 0
    workbook.save(folder_path+main_report_name)

# Function to format and save the interest transactions sheet
def interest_transactions_formatter(interest_transction_df,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Interest Transactions')
    workbook.active = worksheet
    # create a new sheet
    worksheet = workbook.active

    start_row_for_monthly,start_col = 2,2

    interest_transction_list = interest_transction_df["Interest Transactions"]
    if len(interest_transction_list) > 0:
        interest_transaction_height, interest_transaction_width = interest_transction_list[0].shape
    else :
        interest_transaction_height, interest_transaction_width = 0,4
    worksheet.cell(row=start_row_for_monthly, column=start_col, value="Interest Transactions")
    worksheet.cell(row=start_row_for_monthly, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row_for_monthly, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row_for_monthly, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row_for_monthly, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row_for_monthly, start_column=start_col, end_row=start_row_for_monthly, end_column=start_col+interest_transaction_width-1)

    current_row = start_row_for_monthly+1
    month_name_list = []
    curr_year,curr_year_idx,month_start_list = None,None,[]
    for i in range(len(interest_transction_list)):
        interest_transaction_height, interest_transaction_width = interest_transction_list[i].shape
        # Write the column names only for the first time
        if i == 0:
            # Add a row by taking the column names ofmfirst dataframe
            col_name_list = interest_transction_list[i].columns.tolist()
            for c_idx, value in enumerate(col_name_list, start=start_col):
                worksheet.cell(row=current_row, column=c_idx, value=value)
                worksheet.cell(row=current_row, column=c_idx).font = Font(bold=True,color='FFFFFF')
                worksheet.cell(row=current_row, column=c_idx).alignment = Alignment(horizontal='center')
                worksheet.cell(row=current_row, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
                if c_idx == start_col:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                elif c_idx == start_col + interest_transaction_width - 1:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                else:
                    worksheet.cell(row=current_row, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            current_row += 1

        rows = dataframe_to_rows(interest_transction_list[i], index=False, header=False)

        start_row = current_row
        nan_rows = 0

        for r_idx, row in enumerate(rows, start=start_row):
            month_start_list.append(r_idx+1)
            if r_idx == start_row:
                for c_idx, value in enumerate(row, start=start_col):
                    if c_idx == start_col :
                        if curr_year == None:
                            #convert month year to year
                            curr_year = value.split(" ")[1]
                            curr_year_idx = r_idx
                            worksheet.cell(row=r_idx,column=start_col-1,value=curr_year)
                            # fill yellow highlight
                            worksheet.cell(row=r_idx,column=start_col-1).fill= PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        else :
                            if curr_year != value.split(" ")[1]:
                                curr_year = value.split(" ")[1]
                                worksheet.cell(row=r_idx,column=start_col-1,value=curr_year)
                                worksheet.cell(row=r_idx,column=start_col-1).fill= PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                                worksheet.row_dimensions.group(curr_year_idx+1, r_idx-1, hidden=True, outline_level=1)
                                curr_year_idx = r_idx
                    month_name_list.append(r_idx)
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
                    worksheet.cell(row=r_idx, column=c_idx).font = Font(bold=True,color='FFFFFF')
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                    worksheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')

                    if c_idx == start_col:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))
                    elif c_idx == start_col + interest_transaction_width - 1:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                    else:
                        worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thick'))

            for c_idx, value in enumerate(row, start=start_col):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                if c_idx >= start_col + interest_transaction_width - 2:
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '0.00'

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thick'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
                elif c_idx == start_col + interest_transaction_width - 1:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thick'))
                else:
                    worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'))
            current_row += 1

            worksheet.row_dimensions.group(curr_year_idx+1, current_row-1, hidden=True, outline_level=1)

    # Merge all the cells of the month name rows
    for i in range(len(month_name_list)):
        worksheet.cell(row=month_name_list[i], column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=month_name_list[i], column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=month_name_list[i], start_column=start_col, end_row=month_name_list[i], end_column=start_col+interest_transaction_width-1)

    # do auto fit for all columns
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+3)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    # remove gridlines

    worksheet.sheet_view.showGridLines = False
    workbook.active = 0
    workbook.save(folder_path+main_report_name)

# Function to generate the exceptional transactions
def exceptional_transaction(transaction_df):
    data = transaction_df.copy()

    # Get the transactions which are either credits or debits and have amount greater than 1000000
    high_transactions = data.loc[(data['Credit'] > 1000000) | (data['Debit'] > 1000000), ['Date','Description','Credit','Debit']]
    high_transactions.reset_index(drop=True, inplace=True)
    # RTGS less than 2 lakhs
    rtgs_transactions = data.loc[(data['Transaction Type'] == 'rtgs') & (data['Credit'] < 200000), ['Date','Description','Credit','Debit']]
    rtgs_transactions.reset_index(drop=True, inplace=True)
    # Transactions happening on sunday use Day Name column
    sunday_transactions = data.loc[data['Day Name'] == 'Sunday', ['Date','Description','Credit','Debit']]
    sunday_transactions.reset_index(drop=True, inplace=True)

    return high_transactions,rtgs_transactions,sunday_transactions

# Function to format and save the exceptional transactions sheet
def exceptional_transaction_formatter(high_transactions,rtgs_transactions,sunday_transactions,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Exceptional Transactions')
    workbook.active = worksheet
    worksheet = workbook.active


    rows = dataframe_to_rows(high_transactions, index=False, header=False)
    height,width =  high_transactions.shape

    start_row,start_col = 2,2
    # for thick border, slighlty reduce the thickness of the border

    worksheet.cell(row=start_row, column=start_col, value="Summary of Exceptional Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+width-1)

    column_names = high_transactions.columns.to_list()
    curr_row = start_row+1
    for i in range(start_col,width+start_col):
        worksheet.cell(row=start_row+1, column=i, value=column_names[i-start_col])
        worksheet.cell(row=start_row+1, column=i).alignment = Alignment(horizontal='center')
        worksheet.cell(row=start_row+1, column=i).font = Font(bold=True, color='FFFFFF')
        worksheet.cell(row=start_row+1, column=i).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
        worksheet.cell(row=start_row+1, column=i).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    curr_row += 1

    # merge first 2 cells and write Category in that and merge other 2 and write Number in that with grey fill and black font
    worksheet.cell(row=curr_row, column=start_col, value="Category")
    worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

    worksheet.cell(row=curr_row, column=start_col+2, value="Number")
    worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col+2, end_row=curr_row, end_column=start_col+3)

    # Firstly, write for high transactions, make a row with Cateogy as High Transactions and Number as the number of high transactions
    curr_row += 1
    worksheet.cell(row=curr_row, column=start_col, value="High Transactions")
    worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

    worksheet.cell(row=curr_row, column=start_col+2, value=high_transactions.shape[0])
    worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col+2, end_row=curr_row, end_column=start_col+3)

    # Next, write the high transactions
    curr_row += 1
    rows = dataframe_to_rows(high_transactions, index=False, header=False)
    for r_idx, row in enumerate(rows, curr_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # If the height is more than 0, group the rows and collapse them
    if height > 0:
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)

    curr_row += height

    # Next, write for RTGS transactions, make a row with Cateogy as RTGS Transactions and Number as the number of RTGS transactions

    worksheet.cell(row=curr_row, column=start_col, value="RTGS Transactions")
    worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

    worksheet.cell(row=curr_row, column=start_col+2, value=rtgs_transactions.shape[0])
    worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col+2, end_row=curr_row, end_column=start_col+3)

    # Next, write the RTGS transactions
    curr_row += 1
    rows = dataframe_to_rows(rtgs_transactions, index=False, header=False)
    height,width =  rtgs_transactions.shape
    for r_idx, row in enumerate(rows, curr_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    if height > 0:
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
    curr_row += height

    # Next, write for Sunday transactions, make a row with Cateogy as Sunday Transactions and Number as the number of Sunday transactions

    worksheet.cell(row=curr_row, column=start_col, value="Sunday Transactions")
    worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

    worksheet.cell(row=curr_row, column=start_col+2, value=sunday_transactions.shape[0])
    worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col+2, end_row=curr_row, end_column=start_col+3)

    # Next, write the Sunday transactions
    curr_row += 1
    rows = dataframe_to_rows(sunday_transactions, index=False, header=False)
    height,width =  sunday_transactions.shape
    for r_idx, row in enumerate(rows, curr_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    if height > 0:
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
    # Write a row with in Category as Total and Number as the sum of all the transactions
    curr_row += height
    worksheet.cell(row=curr_row, column=start_col, value="Total")
    worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

    worksheet.cell(row=curr_row, column=start_col+2, value=high_transactions.shape[0]+rtgs_transactions.shape[0]+sunday_transactions.shape[0])
    worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True, color='000000')
    worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=curr_row, start_column=start_col+2, end_row=curr_row, end_column=start_col+3)

    # do auto fit for all columns
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+3)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width


    worksheet.sheet_view.showGridLines = False
    workbook.active = 0
    # Save the file
    workbook.save(folder_path +main_report_name)

# Function to generate the inhouse transactions
def inhouse_transactions(transaction_df):
    data = transaction_df.copy()
    inhouse_transactions = data.loc[(data['Related Party'] == 'related party'), ['Date','Description','Credit','Debit','Cleaned Entity Name']]
    inhouse_transactions.reset_index(drop=True, inplace=True)
    inhouse_transactions_grouped = inhouse_transactions.groupby('Cleaned Entity Name').agg({'Credit':'sum','Debit':'sum'}).reset_index()
    inhouse_transactions_grouped['Total'] = inhouse_transactions_grouped['Credit'] + inhouse_transactions_grouped['Debit']
    inhouse_transactions_grouped['No. of Credit Transactions'] = inhouse_transactions[inhouse_transactions['Credit']>0].groupby('Cleaned Entity Name').size().reset_index(name='No. of Credit Transactions')['No. of Credit Transactions']
    inhouse_transactions_grouped['No. of Debit Transactions'] = inhouse_transactions[inhouse_transactions['Debit']>0].groupby('Cleaned Entity Name').size().reset_index(name='No. of Debit Transactions')['No. of Debit Transactions']
    inhouse_transactions_grouped = inhouse_transactions_grouped.sort_values(by='Total', ascending=False)

    return inhouse_transactions,inhouse_transactions_grouped

# Function to format and save the inhouse transactions sheet
def inhouse_transactions_formatter(inhouse_transactions,inhouse_transactions_grouped,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Inhouse Transactions')
    workbook.active = worksheet
    worksheet = workbook.active

    height,width =  inhouse_transactions.shape

    start_row,start_col = 2,2

    worksheet.cell(row=start_row, column=start_col, value="Summary of Inhouse Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+width+2-1)

    # Write Date in B2, Description in C2+D2 merged, Receipts in E2+F2 merges, Payments in G2+H2
    worksheet.cell(row=start_row+1, column=start_col, value="Date")
    worksheet.cell(row=start_row+1, column=start_col).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+1, value="Description")
    worksheet.cell(row=start_row+1, column=start_col+1).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col+1).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+1).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+1).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+3, value="Credit")
    worksheet.cell(row=start_row+1, column=start_col+3).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+3).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+3).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col+3, end_row=start_row+1, end_column=start_col+4)

    worksheet.cell(row=start_row+1, column=start_col+5, value="Debit")
    worksheet.cell(row=start_row+1, column=start_col+5).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+5).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+5).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col+5, end_row=start_row+1, end_column=start_col+6)

    # Write No. in E4, Amount in F4 , No. in G4, Amount in H4 and merge B3 and B4 and C3 and C4

    worksheet.cell(row=start_row+2, column=start_col+3, value="No.")
    worksheet.cell(row=start_row+2, column=start_col+3).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+3).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+3).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+2, column=start_col+4, value="Amount")
    worksheet.cell(row=start_row+2, column=start_col+4).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+4).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+4).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+4).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+2, column=start_col+5, value="No.")
    worksheet.cell(row=start_row+2, column=start_col+5).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+5).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+5).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+2, column=start_col+6, value="Amount")
    worksheet.cell(row=start_row+2, column=start_col+6).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+6).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+6).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+6).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Now merge  B3 B4 and C3 C4
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+2, end_column=start_col)
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col+1, end_row=start_row+2, end_column=start_col+2)

    # Add a row with total amount involved at the beginning
    worksheet.cell(row=start_row+3, column=start_col, value="Total Amount Involved")
    worksheet.cell(row=start_row+3, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+3, column=start_col).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+3, start_column=start_col, end_row=start_row+3, end_column=start_col+2)

    # Write the sum of numbers and amount in the respective columns using inhouse_transactions_grouped
    worksheet.cell(row=start_row+3, column=start_col+3, value=np.sum(inhouse_transactions_grouped["No. of Credit Transactions"]))
    worksheet.cell(row=start_row+3, column=start_col+3).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+3, column=start_col+3).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+3).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+3, column=start_col+4, value=np.sum(inhouse_transactions_grouped["Credit"]))
    worksheet.cell(row=start_row+3, column=start_col+4).number_format = '#,##0.00'
    if pd.isnull(np.sum(inhouse_transactions_grouped["Credit"])):
        worksheet.cell(row=start_row+3, column=start_col+4, value='-')
    worksheet.cell(row=start_row+3, column=start_col+4).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+3, column=start_col+4).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+4).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+4).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+3, column=start_col+5, value=np.sum(inhouse_transactions_grouped["No. of Debit Transactions"]))
    worksheet.cell(row=start_row+3, column=start_col+5).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+3, column=start_col+5).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+5).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+3, column=start_col+6, value=np.sum(inhouse_transactions_grouped["Debit"]))
    worksheet.cell(row=start_row+3, column=start_col+6).number_format = '#,##0.00'
    if pd.isnull(np.sum(inhouse_transactions_grouped["Debit"])):
        worksheet.cell(row=start_row+3, column=start_col+6, value='-')
    worksheet.cell(row=start_row+3, column=start_col+6).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+3, column=start_col+6).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+6).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+6).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Now write the data in the remaining rows
    curr_row = start_row+4
    column_list = [start_col,start_col+1,start_col+4,start_col+6]
    # Iterate throgh the inhouse transaction group, make an entry similar to the total amount involved and then use inhouse_transactions to write the data of that entity
    for index, row in inhouse_transactions_grouped.iterrows():
        worksheet.cell(row=curr_row, column=start_col, value=row['Cleaned Entity Name'])
        worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+2)

        worksheet.cell(row=curr_row, column=start_col+3, value=row['No. of Credit Transactions'])
        worksheet.cell(row=curr_row, column=start_col+3).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col+3).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col+3).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        #if pd.isnull(row['Credit']):
        #    worksheet.cell(row=curr_row, column=start_col+4, value='-')
        #else:
        worksheet.cell(row=curr_row, column=start_col+4, value=row['Credit'])
        worksheet.cell(row=curr_row, column=start_col+4).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+4).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+4).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+4).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+4).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))


        worksheet.cell(row=curr_row, column=start_col+5, value=row['No. of Debit Transactions'])
        worksheet.cell(row=curr_row, column=start_col+5).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col+5).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col+5).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        #if pd.isnull(row['Debit']):
        #    worksheet.cell(row=curr_row, column=start_col+6, value='-')
        #else:
        worksheet.cell(row=curr_row, column=start_col+6, value=row['Debit'])
        worksheet.cell(row=curr_row, column=start_col+6).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+6).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+6).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+6).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+6).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+6).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        curr_row += 1

        # Search for the entity in inhouse_transactions and write the data in the sheet
        entity_data = inhouse_transactions.loc[inhouse_transactions['Cleaned Entity Name'] == row['Cleaned Entity Name'], ['Date','Description','Credit','Debit']]
        entity_data.reset_index(drop=True, inplace=True)
        height,width = entity_data.shape

        rows = dataframe_to_rows(entity_data, index=False, header=False)

        for r_idx, row in enumerate(rows, curr_row):
            for c_idx, value in zip(column_list, row):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                if c_idx == start_col+4 or c_idx == start_col+6:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='right')
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'
                    #if value is nan then make it -
                    if pd.isnull(value):
                        worksheet.cell(row=r_idx, column=c_idx, value='-')

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            # Make border for start_col+3 and start_col+5 as well
            worksheet.cell(row=r_idx, column=start_col+3).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            worksheet.cell(row=r_idx, column=start_col+5).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Merge the cells for Description
            worksheet.merge_cells(start_row=r_idx, start_column=start_col+1, end_row=r_idx, end_column=start_col+2)

        # Group the rows for the entity
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
        curr_row += height

    #Set the column width,th efont size might be different for different columns and rows
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    worksheet.sheet_view.showGridLines = False

    workbook.active = 0
    workbook.save(folder_path + main_report_name)

# Function to generate the salary transactions
def salary_transactions(transaction_df):
    data = transaction_df.copy()
    salary_transactions = data.loc[(data['Purpose'] == 'salary'), ['Date','Description','Debit','Reconciled Name']]
    salary_transactions.reset_index(drop=True, inplace=True)
    salary_transactions_grouped = salary_transactions.groupby('Reconciled Name').agg({'Debit':'sum'}).reset_index()
    salary_transactions_grouped['No. of Transactions'] = salary_transactions.groupby('Reconciled Name').size().reset_index(name='No. of Transactions')['No. of Transactions']
    salary_transactions_grouped = salary_transactions_grouped.sort_values(by='Debit', ascending=False)

    return salary_transactions,salary_transactions_grouped

# Function to format and save the salary transactions sheet
def salary_transactions_formatter(salary_transactions,salary_transactions_grouped,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Salary Transactions')
    workbook.active = worksheet
    worksheet = workbook.active

    height,width =  salary_transactions.shape

    start_row,start_col = 2,2

    worksheet.cell(row=start_row, column=start_col, value="Summary of Salary Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=4)

    worksheet.cell(row=start_row+1, column=start_col, value="Date")
    worksheet.cell(row=start_row+1, column=start_col).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+1, value="Description")
    worksheet.cell(row=start_row+1, column=start_col+1).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col+1).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+1).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+1).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+2, value="Amount")
    worksheet.cell(row=start_row+1, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+2).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+2).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Add a row with total amount involved at the beginning
    worksheet.cell(row=start_row+2, column=start_col, value="Total Amount Involved")
    worksheet.cell(row=start_row+2, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=start_col+1)

    # Write the sum of all salary transactions in the respective columns using salary_transactions_grouped
    worksheet.cell(row=start_row+2, column=start_col+2, value=np.sum(salary_transactions_grouped["Debit"]))
    worksheet.cell(row=start_row+2, column=start_col+2).number_format = '#,##0.00'
    if pd.isnull(np.sum(salary_transactions_grouped["Debit"])):
        worksheet.cell(row=start_row+2, column=start_col+2, value='-')
    worksheet.cell(row=start_row+2, column=start_col+2).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+2, column=start_col+2).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Now write the data in the remaining rows
    curr_row = start_row+3
    column_list = [start_col,start_col+1,start_col+2]
    # Iterate throgh the salary transaction group, make an entry similar to the total amount involved and then use salary_transactions to write the data of that entity
    for index, row in salary_transactions_grouped.iterrows():
        worksheet.cell(row=curr_row, column=start_col, value=row['Reconciled Name'])
        worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

        #if pd.isnull(row['Debit']):
        #    worksheet.cell(row=curr_row, column=start_col+2, value='-')
        #else:
        worksheet.cell(row=curr_row, column=start_col+2, value=row['Debit'])
        worksheet.cell(row=curr_row, column=start_col+2).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        curr_row += 1

        # Search for the entity in salary_transactions and write the data in the sheet
        entity_data = salary_transactions.loc[salary_transactions['Reconciled Name'] == row['Reconciled Name'], ['Date','Description','Debit']]
        entity_data.reset_index(drop=True, inplace=True)
        height,width = entity_data.shape

        rows = dataframe_to_rows(entity_data, index=False, header=False)

        for r_idx, row in enumerate(rows, curr_row):
            for c_idx, value in zip(column_list, row):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                if c_idx == start_col+2:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='right')
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'
                    #if value is nan then make it -
                    if pd.isnull(value):
                        worksheet.cell(row=r_idx, column=c_idx, value='-')

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')

        # Group the rows for the entity
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
        curr_row += height

    #Set the column width,th efont size might be different for different columns and rows
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    worksheet.sheet_view.showGridLines = False

    workbook.active = 0
    workbook.save(folder_path + main_report_name)

# Function to generate the supplier transactions
def supplier_transactions(transaction_df):
    data = transaction_df.copy()
    supplier_transactions = data.loc[(data['Purpose'] == 'supplier'), ['Date','Description','Debit','Reconciled Name']]
    supplier_transactions.reset_index(drop=True, inplace=True)
    supplier_transactions_grouped = supplier_transactions.groupby('Reconciled Name').agg({'Debit':'sum'}).reset_index()
    supplier_transactions_grouped['No. of Transactions'] = supplier_transactions.groupby('Reconciled Name').size().reset_index(name='No. of Transactions')['No. of Transactions']
    supplier_transactions_grouped = supplier_transactions_grouped.sort_values(by='Debit', ascending=False)

    return supplier_transactions,supplier_transactions_grouped

# Function to format and save the supplier transactions sheet
def supplier_transactions_formatter(supplier_transactions,supplier_transactions_grouped,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Supplier Transactions')
    workbook.active = worksheet
    worksheet = workbook.active

    height,width =  supplier_transactions.shape

    start_row,start_col = 2,2

    worksheet.cell(row=start_row, column=start_col, value="Summary of Supplier Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=4)

    worksheet.cell(row=start_row+1, column=start_col, value="Date")
    worksheet.cell(row=start_row+1, column=start_col).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+1, value="Description")
    worksheet.cell(row=start_row+1, column=start_col+1).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col+1).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+1).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+1).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+2, value="Amount")
    worksheet.cell(row=start_row+1, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+2).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+2).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Add a row with total amount involved at the beginning
    worksheet.cell(row=start_row+2, column=start_col, value="Total Amount Involved")
    worksheet.cell(row=start_row+2, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=start_col+1)

    # Write the sum of all salary transactions in the respective columns using salary_transactions_grouped
    worksheet.cell(row=start_row+2, column=start_col+2, value=np.sum(supplier_transactions_grouped["Debit"]))
    worksheet.cell(row=start_row+2, column=start_col+2).number_format = '#,##0.00'
    if pd.isnull(np.sum(supplier_transactions_grouped["Debit"])):
        worksheet.cell(row=start_row+2, column=start_col+2, value='-')
    worksheet.cell(row=start_row+2, column=start_col+2).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+2, column=start_col+2).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Now write the data in the remaining rows
    curr_row = start_row+3
    column_list = [start_col,start_col+1,start_col+2]
    # Iterate throgh the salary transaction group, make an entry similar to the total amount involved and then use salary_transactions to write the data of that entity
    for index, row in supplier_transactions_grouped.iterrows():
        worksheet.cell(row=curr_row, column=start_col, value=row['Reconciled Name'])
        worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

        worksheet.cell(row=curr_row, column=start_col+2, value=row['Debit'])
        worksheet.cell(row=curr_row, column=start_col+2).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        curr_row += 1

        # Search for the entity in salary_transactions and write the data in the sheet
        entity_data = supplier_transactions.loc[supplier_transactions['Reconciled Name'] == row['Reconciled Name'], ['Date','Description','Debit']]
        entity_data.reset_index(drop=True, inplace=True)
        height,width = entity_data.shape

        rows = dataframe_to_rows(entity_data, index=False, header=False)

        for r_idx, row in enumerate(rows, curr_row):
            for c_idx, value in zip(column_list, row):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                if c_idx == start_col+2:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='right')
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'
                    #if value is nan then make it -
                    if pd.isnull(value):
                        worksheet.cell(row=r_idx, column=c_idx, value='-')

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')

        # Group the rows for the entity
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
        curr_row += height

    #Set the column width,th efont size might be different for different columns and rows
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    worksheet.sheet_view.showGridLines = False

    workbook.active = 0
    workbook.save(folder_path + main_report_name)

# Function to generate the customer transactions
def customer_transactions(transaction_df):
    data = transaction_df.copy()
    customer_transactions = data.loc[(data['Purpose'] == 'customer'), ['Date','Description','Credit','Reconciled Name']]
    customer_transactions.reset_index(drop=True, inplace=True)
    customer_transactions_grouped = customer_transactions.groupby('Reconciled Name').agg({'Credit':'sum'}).reset_index()
    customer_transactions_grouped['No. of Transactions'] = customer_transactions.groupby('Reconciled Name').size().reset_index(name='No. of Transactions')['No. of Transactions']
    customer_transactions_grouped = customer_transactions_grouped.sort_values(by='Credit', ascending=False)

    return customer_transactions,customer_transactions_grouped

# Function to format and save the customer transactions sheet
def customer_transactions_formatter(customer_transactions,customer_transactions_grouped,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Customer Transactions')
    workbook.active = worksheet
    worksheet = workbook.active

    height,width =  customer_transactions.shape

    start_row,start_col = 2,2

    worksheet.cell(row=start_row, column=start_col, value="Summary of Customer Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=4)

    worksheet.cell(row=start_row+1, column=start_col, value="Date")
    worksheet.cell(row=start_row+1, column=start_col).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+1, value="Description")
    worksheet.cell(row=start_row+1, column=start_col+1).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col+1).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+1).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+1).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+2, value="Amount")
    worksheet.cell(row=start_row+1, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+2).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+2).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Add a row with total amount involved at the beginning
    worksheet.cell(row=start_row+2, column=start_col, value="Total Amount Involved")
    worksheet.cell(row=start_row+2, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=start_col+1)

    # Write the sum of all salary transactions in the respective columns using salary_transactions_grouped
    worksheet.cell(row=start_row+2, column=start_col+2, value=np.sum(customer_transactions_grouped["Credit"]))
    worksheet.cell(row=start_row+2, column=start_col+2).number_format = '#,##0.00'
    if pd.isnull(np.sum(customer_transactions_grouped["Credit"])):
        worksheet.cell(row=start_row+2, column=start_col+2, value='-')
    worksheet.cell(row=start_row+2, column=start_col+2).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+2, column=start_col+2).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Now write the data in the remaining rows
    curr_row = start_row+3
    column_list = [start_col,start_col+1,start_col+2]
    # Iterate throgh the salary transaction group, make an entry similar to the total amount involved and then use salary_transactions to write the data of that entity
    for index, row in customer_transactions_grouped.iterrows():
        worksheet.cell(row=curr_row, column=start_col, value=row['Reconciled Name'])
        worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

        worksheet.cell(row=curr_row, column=start_col+2, value=row['Credit'])
        worksheet.cell(row=curr_row, column=start_col+2).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        curr_row += 1

        # Search for the entity in salary_transactions and write the data in the sheet
        entity_data = customer_transactions.loc[customer_transactions['Reconciled Name'] == row['Reconciled Name'], ['Date','Description','Credit']]
        entity_data.reset_index(drop=True, inplace=True)
        height,width = entity_data.shape

        rows = dataframe_to_rows(entity_data, index=False, header=False)

        for r_idx, row in enumerate(rows, curr_row):
            for c_idx, value in zip(column_list, row):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                if c_idx == start_col+2:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='right')
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'
                    #if value is nan then make it -
                    if pd.isnull(value):
                        worksheet.cell(row=r_idx, column=c_idx, value='-')

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')

        # Group the rows for the entity
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
        curr_row += height

    #Set the column width,th efont size might be different for different columns and rows
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    worksheet.sheet_view.showGridLines = False

    workbook.active = 0
    workbook.save(folder_path + main_report_name)

# Function to generate the charges transactions
def charges_transactions(transaction_df):
    data = transaction_df.copy()
    charges_transactions = data.loc[(data['Transaction Type'] == 'charges'), ['Date','Description','Debit','Charge Type']]
    charges_transactions.reset_index(drop=True, inplace=True)
    charges_transactions_grouped = charges_transactions.groupby('Charge Type').agg({'Debit':'sum'}).reset_index()
    charges_transactions_grouped['No. of Transactions'] = charges_transactions.groupby('Charge Type').size().reset_index(name='No. of Transactions')['No. of Transactions']
    charges_transactions_grouped = charges_transactions_grouped.sort_values(by='Debit', ascending=False)

    return charges_transactions,charges_transactions_grouped

# Function to format and save the charges transactions sheet
def charges_transactions_formatter(charges_transactions,charges_transactions_grouped,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Charges Transactions')
    workbook.active = worksheet
    worksheet = workbook.active

    height,width =  charges_transactions.shape

    start_row,start_col = 2,2

    worksheet.cell(row=start_row, column=start_col, value="Summary of Charges Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=4)

    worksheet.cell(row=start_row+1, column=start_col, value="Date")
    worksheet.cell(row=start_row+1, column=start_col).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+1, value="Description")
    worksheet.cell(row=start_row+1, column=start_col+1).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col+1).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+1).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+1).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+2, value="Amount")
    worksheet.cell(row=start_row+1, column=start_col+2).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+2).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+2).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Add a row with total amount involved at the beginning
    worksheet.cell(row=start_row+2, column=start_col, value="Total Amount Involved")
    worksheet.cell(row=start_row+2, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=start_col+1)

    # Write the sum of all salary transactions in the respective columns using salary_transactions_grouped
    worksheet.cell(row=start_row+2, column=start_col+2, value=np.sum(charges_transactions_grouped["Debit"]))
    worksheet.cell(row=start_row+2, column=start_col+2).number_format = '#,##0.00'
    if pd.isnull(np.sum(charges_transactions_grouped["Debit"])):
        worksheet.cell(row=start_row+2, column=start_col+2, value='-')
    worksheet.cell(row=start_row+2, column=start_col+2).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+2, column=start_col+2).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Now write the data in the remaining rows
    curr_row = start_row+3
    column_list = [start_col,start_col+1,start_col+2]
    # Iterate throgh the salary transaction group, make an entry similar to the total amount involved and then use salary_transactions to write the data of that entity
    for index, row in charges_transactions_grouped.iterrows():
        worksheet.cell(row=curr_row, column=start_col, value=row['Charge Type'])
        worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)

        worksheet.cell(row=curr_row, column=start_col+2, value=row['Debit'])
        worksheet.cell(row=curr_row, column=start_col+2).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+2).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+2).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+2).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+2).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        curr_row += 1

        # Search for the entity in salary_transactions and write the data in the sheet
        entity_data = charges_transactions.loc[charges_transactions['Charge Type'] == row['Charge Type'], ['Date','Description','Debit']]
        entity_data.reset_index(drop=True, inplace=True)
        height,width = entity_data.shape

        rows = dataframe_to_rows(entity_data, index=False, header=False)

        for r_idx, row in enumerate(rows, curr_row):
            for c_idx, value in zip(column_list, row):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                if c_idx == start_col+2:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='right')
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'
                    #if value is nan then make it -
                    if pd.isnull(value):
                        worksheet.cell(row=r_idx, column=c_idx, value='-')

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')

        # Group the rows for the entity
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
        curr_row += height

    #Set the column width,th efont size might be different for different columns and rows
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    worksheet.sheet_view.showGridLines = False

    workbook.active = 0
    workbook.save(folder_path + main_report_name)

# Function to generate the return transactions
def transaction_return(transaction_df):
    data = transaction_df.copy()
    # Go through each row and see if the transaction is either cheque return, online return or ecs return and assign it to Type of return
    for index, row in data.iterrows():
        if row["Cheque Return"] == "yes":
            data.at[index, 'Type of return'] = 'Cheque Return'

        elif row["Online Return"] == "yes":
            data.at[index, 'Type of return'] = 'Online Return'
        elif row["ECS Return"] == "yes":
            data.at[index, 'Type of return'] = 'ECS Return'
        else:
            data.at[index, 'Type of return'] = 'Not Applicable'

    # transaction return should not have Not applicable
    transaction_return = data.loc[(data['Type of return'] != 'Not Applicable'), ['Date','Description','Credit','Debit','Type of return']]
    transaction_return.reset_index(drop=True, inplace=True)
    transaction_return_grouped = transaction_return.groupby('Type of return').agg({'Credit':'sum','Debit':'sum'}).reset_index()
    transaction_return_grouped['Total'] = transaction_return_grouped['Credit'] + transaction_return_grouped['Debit']
    transaction_return_grouped['No. of Credit Transactions'] = transaction_return[transaction_return['Credit']>0].groupby('Type of return').size().reset_index(name='No. of Credit Transactions')['No. of Credit Transactions']
    transaction_return_grouped['No. of Debit Transactions'] = transaction_return[transaction_return['Debit']>0].groupby('Type of return').size().reset_index(name='No. of Debit Transactions')['No. of Debit Transactions']
    transaction_return_grouped = transaction_return_grouped.sort_values(by='Total', ascending=False)


    return transaction_return,transaction_return_grouped

# Function to format and save the return transactions sheet
def return_transactions_formatter(inhouse_transactions,inhouse_transactions_grouped,folder_path,main_report_name):
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='Return Transactions')
    workbook.active = worksheet
    # create a new sheet
    worksheet = workbook.active
    height,width =  inhouse_transactions.shape

    start_row,start_col = 2,2

    worksheet.cell(row=start_row, column=start_col, value="Summary of Return Transactions")
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True,underline='single', color='FFFFFF', size=12)
    worksheet.cell(row=start_row, column=start_col).fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    worksheet.cell(row=start_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+width+2-1)

    # Write Date in B2, Description in C2+D2 merged, Receipts in E2+F2 merges, Payments in G2+H2
    worksheet.cell(row=start_row+1, column=start_col, value="Date")
    worksheet.cell(row=start_row+1, column=start_col).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+1, value="Description")
    worksheet.cell(row=start_row+1, column=start_col+1).alignment = Alignment(horizontal='center',vertical='center')
    worksheet.cell(row=start_row+1, column=start_col+1).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+1).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+1).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+1, column=start_col+3, value="Credit")
    worksheet.cell(row=start_row+1, column=start_col+3).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+3).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+3).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col+3, end_row=start_row+1, end_column=start_col+4)

    worksheet.cell(row=start_row+1, column=start_col+5, value="Debit")
    worksheet.cell(row=start_row+1, column=start_col+5).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+1, column=start_col+5).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+1, column=start_col+5).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+1, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col+5, end_row=start_row+1, end_column=start_col+6)

    # Write No. in E4, Amount in F4 , No. in G4, Amount in H4 and merge B3 and B4 and C3 and C4

    worksheet.cell(row=start_row+2, column=start_col+3, value="No.")
    worksheet.cell(row=start_row+2, column=start_col+3).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+3).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+3).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+2, column=start_col+4, value="Amount")
    worksheet.cell(row=start_row+2, column=start_col+4).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+4).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+4).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+4).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+2, column=start_col+5, value="No.")
    worksheet.cell(row=start_row+2, column=start_col+5).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+5).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+5).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+2, column=start_col+6, value="Amount")
    worksheet.cell(row=start_row+2, column=start_col+6).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+2, column=start_col+6).font = Font(bold=True, color='FFFFFF',size=11.5)
    worksheet.cell(row=start_row+2, column=start_col+6).fill = PatternFill(start_color='6CC9E5', end_color='6CC9E5', fill_type='solid')
    worksheet.cell(row=start_row+2, column=start_col+6).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Now merge  B3 B4 and C3 C4
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+2, end_column=start_col)
    worksheet.merge_cells(start_row=start_row+1, start_column=start_col+1, end_row=start_row+2, end_column=start_col+2)

    # Add a row with total amount involved at the beginning
    worksheet.cell(row=start_row+3, column=start_col, value="Total Amount Involved")
    worksheet.cell(row=start_row+3, column=start_col).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+3, column=start_col).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    worksheet.merge_cells(start_row=start_row+3, start_column=start_col, end_row=start_row+3, end_column=start_col+2)

    # Write the sum of numbers and amount in the respective columns using inhouse_transactions_grouped
    worksheet.cell(row=start_row+3, column=start_col+3, value=np.sum(inhouse_transactions_grouped["No. of Credit Transactions"]))
    worksheet.cell(row=start_row+3, column=start_col+3).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+3, column=start_col+3).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+3).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+3, column=start_col+4, value=np.sum(inhouse_transactions_grouped["Credit"]))
    worksheet.cell(row=start_row+3, column=start_col+4).number_format = '#,##0.00'
    if pd.isnull(np.sum(inhouse_transactions_grouped["Credit"])):
        worksheet.cell(row=start_row+3, column=start_col+4, value='-')
    worksheet.cell(row=start_row+3, column=start_col+4).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+3, column=start_col+4).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+4).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+4).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+3, column=start_col+5, value=np.sum(inhouse_transactions_grouped["No. of Debit Transactions"]))
    worksheet.cell(row=start_row+3, column=start_col+5).alignment = Alignment(horizontal='center')
    worksheet.cell(row=start_row+3, column=start_col+5).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+5).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    worksheet.cell(row=start_row+3, column=start_col+6, value=np.sum(inhouse_transactions_grouped["Debit"]))
    worksheet.cell(row=start_row+3, column=start_col+6).number_format = '#,##0.00'
    if pd.isnull(np.sum(inhouse_transactions_grouped["Debit"])):
        worksheet.cell(row=start_row+3, column=start_col+6, value='-')
    worksheet.cell(row=start_row+3, column=start_col+6).alignment = Alignment(horizontal='right')
    worksheet.cell(row=start_row+3, column=start_col+6).font = Font(bold=True,size=11.5)
    worksheet.cell(row=start_row+3, column=start_col+6).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    worksheet.cell(row=start_row+3, column=start_col+6).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    # Now write the data in the remaining rows
    curr_row = start_row+4
    column_list = [start_col,start_col+1,start_col+4,start_col+6]
    # Iterate throgh the inhouse transaction group, make an entry similar to the total amount involved and then use inhouse_transactions to write the data of that entity
    for index, row in inhouse_transactions_grouped.iterrows():
        worksheet.cell(row=curr_row, column=start_col, value=row['Type of return'])
        worksheet.cell(row=curr_row, column=start_col).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        worksheet.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+2)

        worksheet.cell(row=curr_row, column=start_col+3, value=row['No. of Credit Transactions'])
        worksheet.cell(row=curr_row, column=start_col+3).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col+3).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col+3).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+3).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        #if pd.isnull(row['Credit']):
        #    worksheet.cell(row=curr_row, column=start_col+4, value='-')
        #else:
        worksheet.cell(row=curr_row, column=start_col+4, value=row['Credit'])
        worksheet.cell(row=curr_row, column=start_col+4).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+4).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+4).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+4).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+4).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))


        worksheet.cell(row=curr_row, column=start_col+5, value=row['No. of Debit Transactions'])
        worksheet.cell(row=curr_row, column=start_col+5).alignment = Alignment(horizontal='center')
        worksheet.cell(row=curr_row, column=start_col+5).font = Font(bold=True)
        worksheet.cell(row=curr_row, column=start_col+5).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+5).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        #if pd.isnull(row['Debit']):
        #    worksheet.cell(row=curr_row, column=start_col+6, value='-')
        #else:
        worksheet.cell(row=curr_row, column=start_col+6, value=row['Debit'])
        worksheet.cell(row=curr_row, column=start_col+6).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+6).alignment = Alignment(horizontal='right')
        worksheet.cell(row=curr_row, column=start_col+6).number_format = '#,##0.00'
        worksheet.cell(row=curr_row, column=start_col+6).font = Font(bold=True,italic=True)
        worksheet.cell(row=curr_row, column=start_col+6).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        worksheet.cell(row=curr_row, column=start_col+6).border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

        curr_row += 1

        # Search for the entity in inhouse_transactions and write the data in the sheet
        entity_data = inhouse_transactions.loc[inhouse_transactions['Type of return'] == row['Type of return'], ['Date','Description','Credit','Debit']]
        entity_data.reset_index(drop=True, inplace=True)
        height,width = entity_data.shape

        rows = dataframe_to_rows(entity_data, index=False, header=False)

        for r_idx, row in enumerate(rows, curr_row):
            for c_idx, value in zip(column_list, row):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
                worksheet.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                if c_idx == start_col+4 or c_idx == start_col+6:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='right')
                    worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'
                    #if value is nan then make it -
                    if pd.isnull(value):
                        worksheet.cell(row=r_idx, column=c_idx, value='-')

                if c_idx == start_col:
                    worksheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            # Make border for start_col+3 and start_col+5 as well
            worksheet.cell(row=r_idx, column=start_col+3).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            worksheet.cell(row=r_idx, column=start_col+5).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Merge the cells for Description
            worksheet.merge_cells(start_row=r_idx, start_column=start_col+1, end_row=r_idx, end_column=start_col+2)

        # Group the rows for the entity
        worksheet.row_dimensions.group(curr_row, curr_row+height-1, hidden=True)
        curr_row += height

    #Set the column width,th efont size might be different for different columns and rows
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        worksheet.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width

    worksheet.sheet_view.showGridLines = False

    workbook.active = 0
    workbook.save(folder_path + main_report_name)

# Function to generate, format and save the graphs in the main report
def graph_generator(info_extracted_df,folder_path,main_report_name):
    # Get the total balance for each month and sums of credit and debit
    total_balance = info_extracted_df.groupby(['Year', 'Month']).agg({'Month Year':'last','Balance':'last',"Credit":'sum',"Debit":'sum'}).reset_index()

    # Only keep the columns 'Month Year' and 'Balance'
    total_balance = total_balance[['Month Year','Balance','Credit','Debit']]

    # Make a new table with Sum of Debits and Credits for each transaction type
    transaction_type_table = info_extracted_df.groupby(['Transaction Type']).agg({'Month Year':'last','Balance':'last',"Credit":'sum',"Debit":'sum'}).reset_index()
    transaction_type_table = transaction_type_table[['Transaction Type','Credit','Debit']]
    transaction_type_table['Total'] = transaction_type_table['Credit'] + transaction_type_table['Debit']
    #Sort the table in descending order of total amount
    transaction_type_table = transaction_type_table.sort_values(by='Total', ascending=False)

    related_party_table = info_extracted_df.groupby(['Related Party']).agg({'Month Year':'last',"Credit":'sum',"Debit":'sum'}).reset_index()
    related_party_table = related_party_table[['Related Party','Credit','Debit']]
    # Capitalize the first letter of each word in the Related Party column
    related_party_table['Related Party'] = related_party_table['Related Party'].str.title()

    supplier_table = info_extracted_df[info_extracted_df['Purpose'] == 'supplier']
    supplier_table = supplier_table.groupby(['Year','Month']).agg({'Month Year':'last','Debit':'sum'}).reset_index()
    supplier_table = supplier_table[['Month Year','Debit']]

    customer_table = info_extracted_df[info_extracted_df['Purpose'] == 'customer']
    customer_table = customer_table.groupby(['Year','Month']).agg({'Month Year':'last','Credit':'sum'}).reset_index()
    customer_table = customer_table[['Month Year','Credit']]

    salary_table = info_extracted_df[info_extracted_df['Purpose'] == 'salary']
    salary_table = salary_table.groupby(['Year','Month']).agg({'Month Year':'last','Debit':'sum'}).reset_index()
    salary_table = salary_table[['Month Year','Debit']]

    # Create a workbook
    main_report_path = folder_path+main_report_name
    workbook = load_workbook(main_report_path)
    worksheet = workbook.create_sheet(title='All Graphs')
    workbook.active = worksheet
    worksheet = workbook.active
    start_row,start_col = 1,1

    # Add data to the worksheet
    row = dataframe_to_rows(total_balance, index=False, header=True)
    for r_idx, row in enumerate(row, start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).font = Font(color='FFFFFF')
            if c_idx > start_col:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

    # Now create a line chart
    chart = openpyxl.chart.LineChart()
    chart.title,chart.style,chart.y_axis.title,chart.x_axis.title = "Total Balance Graph",13,"Ending Balance","Month-Year"
    data = openpyxl.chart.Reference(worksheet, min_col=2, min_row=1, max_row=total_balance.shape[0]+1, max_col=2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(openpyxl.chart.Reference(worksheet, min_col=1, min_row=2, max_row=total_balance.shape[0]+1))
    chart.shape,chart.height,chart.width = 5,10,20
    s1 = chart.series[0]
    s1.graphicalProperties.line.solidFill = "2CA1F0"
    marker = Marker()
    marker.symbol,marker.graphicalProperties.solidFill,marker.size,marker.graphicalProperties.line.noFill = "circle","1567D1",3,True
    s1.marker,s1.graphicalProperties.line.width = marker,20000
    chart.legend = None
    worksheet.add_chart(chart, "B2")

    # Add new chart with 2 lines for credit and debit sums
    chart = openpyxl.chart.LineChart()
    chart.title,chart.style,chart.y_axis.title,chart.x_axis.title = "Total Credits and Debits Graph",13,"Amount","Month-Year"
    data = openpyxl.chart.Reference(worksheet, min_col=3, min_row=1, max_row=total_balance.shape[0]+1, max_col=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(openpyxl.chart.Reference(worksheet, min_col=1, min_row=2, max_row=total_balance.shape[0]+1))
    chart.shape,chart.height,chart.width = 5,10,20
    s1 = chart.series[0]
    s1.graphicalProperties.line.solidFill = "2CA1F0"
    marker = Marker()
    marker.symbol,marker.graphicalProperties.solidFill,marker.size,marker.graphicalProperties.line.noFill = "circle","1567D1",3,True
    s1.marker,s1.graphicalProperties.line.width = marker,20000

    s2 = chart.series[1]
    s2.graphicalProperties.line.solidFill = "FF0000"
    marker = Marker()
    marker.symbol,marker.graphicalProperties.solidFill,marker.size,marker.graphicalProperties.line.noFill = "circle","FF0000",3,True
    s2.marker,s2.graphicalProperties.line.width = marker,20000

    worksheet.add_chart(chart, "N2")

    # Add data to the worksheet
    start_row,start_col = 1,6
    row = dataframe_to_rows(transaction_type_table, index=False, header=True)
    for r_idx, row in enumerate(row, start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).font = Font(color='FFFFFF')
            if c_idx > start_col:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

    chart = BarChart()
    chart.title,chart.style,chart.y_axis.title,chart.x_axis.title = "Transaction Type Graph",13,"Amount","Transaction Type"

    data = openpyxl.chart.Reference(worksheet, min_col=start_col+1, min_row=1, max_row=transaction_type_table.shape[0]+1, max_col=start_col+2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(openpyxl.chart.Reference(worksheet, min_col=start_col, min_row=2, max_row=transaction_type_table.shape[0]+1))
    chart.shape,chart.height,chart.width = 5,10,20
    s1 = chart.series[0]
    s1.graphicalProperties.solidFill = "2CA1F0"
    s2 = chart.series[1]
    s2.graphicalProperties.solidFill = "FF0000"
    # Set the title and labels
    chart.title = "Transaction Type Graph"
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Transaction Type"
    data_labels = DataLabelList()
    data_labels.showVal = True  # Show the value
    # Add the chart to the worksheet
    worksheet.add_chart(chart, "B25")

    # Add data to the worksheet
    start_row,start_col = 1,11
    row = dataframe_to_rows(related_party_table, index=False, header=True)
    for r_idx, row in enumerate(row, start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).font = Font(color='FFFFFF')
            if c_idx > start_col:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

    # Make a Doughnut Chart with double ring
    chart = openpyxl.chart.DoughnutChart()
    chart.title,chart.style = "Related Party Graph Inner: Credit, Outer: Debit",13
    legend_names = ['Not Related Party', 'Related Party']

    data = openpyxl.chart.Reference(worksheet, min_col=start_col+1, min_row=1, max_row=related_party_table.shape[0]+1, max_col=start_col+2)
    chart.add_data(data, titles_from_data=True)
    chart.shape,chart.height,chart.width = 5,10,20
    chart.set_categories(openpyxl.chart.Reference(worksheet, min_col=start_col, min_row=2, max_row=related_party_table.shape[0]+1))

    # Sert the legend values as not related party, related party
    chart.varyColors = True
    chart.holeSize = 50
    chart.firstSliceAng = 270
    chart.secondPieSize = 50
    chart.secondSliceAng = 270

    # Show the data labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = True
    worksheet.add_chart(chart, "N25")

    # Add data to the worksheet
    start_row,start_col = 1,15
    row = dataframe_to_rows(supplier_table, index=False, header=True)
    for r_idx, row in enumerate(row, start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).font = Font(color='FFFFFF')
            if c_idx > start_col:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

    # Create a line chart for supplier
    chart = openpyxl.chart.LineChart()
    chart.title,chart.style,chart.y_axis.title,chart.x_axis.title = "Supplier Transactions Graph",13,"Amount","Month-Year"
    data = openpyxl.chart.Reference(worksheet, min_col=start_col+1, min_row=1, max_row=supplier_table.shape[0]+1, max_col=start_col+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(openpyxl.chart.Reference(worksheet, min_col=start_col, min_row=2, max_row=supplier_table.shape[0]+1))
    chart.shape,chart.height,chart.width = 5,10,20
    s1 = chart.series[0]
    s1.graphicalProperties.line.solidFill = "2CA1F0"
    marker = Marker()
    marker.symbol,marker.graphicalProperties.solidFill,marker.size,marker.graphicalProperties.line.noFill = "circle","1567D1",3,True
    s1.marker,s1.graphicalProperties.line.width = marker,20000
    chart.legend = None
    worksheet.add_chart(chart, "B48")

    # Add data to the worksheet
    start_row,start_col = 1,19
    row = dataframe_to_rows(customer_table, index=False, header=True)
    for r_idx, row in enumerate(row, start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).font = Font(color='FFFFFF')
            if c_idx > start_col:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

    # Create a line chart for customer
    chart = openpyxl.chart.LineChart()
    chart.title,chart.style,chart.y_axis.title,chart.x_axis.title = "Customer Transactions Graph",13,"Amount","Month-Year"
    data = openpyxl.chart.Reference(worksheet, min_col=start_col+1, min_row=1, max_row=customer_table.shape[0]+1, max_col=start_col+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(openpyxl.chart.Reference(worksheet, min_col=start_col, min_row=2, max_row=customer_table.shape[0]+1))
    chart.shape,chart.height,chart.width = 5,10,20
    s1 = chart.series[0]
    s1.graphicalProperties.line.solidFill = "2CA1F0"
    marker = Marker()
    marker.symbol,marker.graphicalProperties.solidFill,marker.size,marker.graphicalProperties.line.noFill = "circle","1567D1",3,True
    s1.marker,s1.graphicalProperties.line.width = marker,20000
    chart.legend = None
    worksheet.add_chart(chart, "N48")

    # Add data to the worksheet
    start_row,start_col = 1,23
    row = dataframe_to_rows(salary_table, index=False, header=True)
    for r_idx, row in enumerate(row, start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
            worksheet.cell(row=r_idx, column=c_idx).font = Font(color='FFFFFF')
            if c_idx > start_col:
                worksheet.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

    # Create a line chart for salary
    chart = openpyxl.chart.LineChart()
    chart.title,chart.style,chart.y_axis.title,chart.x_axis.title = "Salary Transactions Graph",13,"Amount","Month-Year"
    data = openpyxl.chart.Reference(worksheet, min_col=start_col+1, min_row=1, max_row=salary_table.shape[0]+1, max_col=start_col+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(openpyxl.chart.Reference(worksheet, min_col=start_col, min_row=2, max_row=salary_table.shape[0]+1))
    chart.shape,chart.height,chart.width = 5,10,20
    s1 = chart.series[0]
    s1.graphicalProperties.line.solidFill = "2CA1F0"
    marker = Marker()
    marker.symbol,marker.graphicalProperties.solidFill,marker.size,marker.graphicalProperties.line.noFill = "circle","1567D1",3,True
    s1.marker,s1.graphicalProperties.line.width = marker,20000
    chart.legend = None
    worksheet.add_chart(chart, "B71")
    # Save the workbook
    worksheet.sheet_view.showGridLines = False
    # Worksheet Names
    workbook.active = 0
    workbook.save(folder_path + main_report_name)

# Function to save the dataframes to json files
def save_to_json(folder_path,dataframes):
    # Save the dataframe files to json use the name of dataframes as the name of the json files

    for dataframe in dataframes:
        # store the json name as the variable name of the dataframe
        json_name = dataframe.name
        # Save the dataframe to json
        dataframe.to_json(folder_path +f"\{json_name}"+'.json',orient='records')

    return

#-----------------------------------------The End-----------------------------------------#
