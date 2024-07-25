from multiprocessing.util import info
import pandas as pd
import numpy as np
from report_functions import *
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill,Border,Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.marker import Marker

# Main function to generate the report from the list of csv files and statement details
def report_generator(sheet_list,statement_details,party_name,folder_path,use_ner=False):
    print("Starting Report Generation")
    # Consolidate all the sheets into one
    sheet = sheet_consolidator(sheet_list) 

    # Main excel sheet with all the data columns
    info_extracted_df = description_info_extraction(sheet,party_name,use_ner_model=use_ner)

    # Extract the details from the statement_details
    from_date, to_date, ifsc_list, bank_list, account_type_list = sheet_details_processor(statement_details)
    
    # update statement_details by removing the column File Path
    statement_details = statement_details.drop(columns=['File Path'])
    statement_details.to_excel(os.path.join(folder_path,"statement_details.xlsx"), index=False)
    
    # make dummy values for now
    # from_date = pd.to_datetime('2021-01-01')
    # to_date = pd.to_datetime('2021-12-31')
    # ifsc_list = ['ICIC0000001']
    # bank_list = ['ICICI Bank']
    # account_type_list = ['Savings Account']

    # Save the info_extracted_df to an excel file
    #info_extracted_df.to_excel(folder_path+r"\info_extracted_df.xlsx", index=False)
    info_extracted_df.to_excel(os.path.join(folder_path,"info_extracted_df.xlsx"), index=False)
    # Opening and closing balance
    opening_balance = info_extracted_df['Balance'].iloc[0]
    closing_balance = info_extracted_df['Balance'].iloc[-1]

    # Total months from from_date to to_date
    total_months = (pd.to_datetime(to_date).year - pd.to_datetime(from_date).year)*12 + (pd.to_datetime(to_date).month - pd.to_datetime(from_date).month) + 1   
    
    # Make a workbook object - Main Report
    #main_report_name = r"\Main_Report.xlsx"
    main_report_name = "Main_Report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Landing Page" # type: ignore
    
    ws['G6'] = "Bank Statement Analysis" # type: ignore
    cell = ws['G6']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, underline='single', color='FFFFFF', size=14)
    cell.fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    ws.merge_cells('G6:L6')

    for i in range(7,13):
        for j in range(7,12):
            cell = ws.cell(row=i,column=j)
            cell.fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['G7'] = "Party Name"
    cell = ws['G7']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G7:I7')

    ws['J7'] = party_name
    cell = ws['J7']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J7:L7')

    ws['G8'] = "Start of the Period"
    cell = ws['G8']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G8:I8')

    ws['J8'] = from_date
    cell = ws['J8']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J8:L8')

    ws['G9'] = "End of the Period"
    cell = ws['G9']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G9:I9')

    ws['J9'] = to_date
    cell = ws['J9']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J9:L9')

    ws['G10'] = "Total Months"
    cell = ws['G10']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G10:I10')

    ws['J10'] = total_months
    cell = ws['J10']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J10:L10')

    ws['G11'] = "Opening Balance"
    cell = ws['G11']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G11:I11')

    ws['J11'] = opening_balance
    cell = ws['J11']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    ws.merge_cells('J11:L11')

    ws['G12'] = "Closing Balance"
    cell = ws['G12']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G12:I12')

    ws['J12'] = closing_balance
    cell = ws['J12']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    ws.merge_cells('J12:L12')
    ws.sheet_view.showGridLines = False
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        ws.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width
    wb.active = 0
    wb.save(os.path.join(folder_path,main_report_name))

    statement_details.name = "statement_details"
    save_to_json(folder_path,[statement_details])
    statement_details_formatter(statement_details,folder_path,main_report_name)
    # Adding graphs to the main report
    graph_generator(info_extracted_df,folder_path,main_report_name)

    # name a directory for json files
    json_dir = os.path.join(folder_path,"json_files")
    try:
        os.makedirs(json_dir,exist_ok=True)
    except:
        pass

    info_extracted_df.name = "info_extracted_df"
    save_to_json(json_dir,[info_extracted_df])

    # Adding overview sheet to the main report
    overview_sheet = overview_generator(info_extracted_df)
    overview_sheet.name = "overview_sheet"
    save_to_json(json_dir,[overview_sheet])
    overview_sheet_formatter(overview_sheet,folder_path,main_report_name)

    # Adding summary sheet to the main report
    summary_sheet = summary_of_debits_and_credits(info_extracted_df)
    summary_sheet.name = "summary_sheet"
    save_to_json(json_dir,[summary_sheet])
    summary_sheet_formatter(summary_sheet,folder_path,main_report_name)

    # Adding top credits and debits sheet to the main report
    top_credits,top_debits,top_credits_debits_monthwise = top_credits_and_debits(info_extracted_df)
    top_credits.name = "top_credits"
    top_debits.name = "top_debits"
    top_credits_debits_monthwise.name = "top_credits_debits_monthwise"
    save_to_json(json_dir,[top_credits,top_debits,top_credits_debits_monthwise])
    top_credit_debit_formatter(top_credits,top_debits,top_credits_debits_monthwise,folder_path,main_report_name)

    # Adding exceptional transactions sheet to the main report
    high_transactions,rtgs_transactions,sunday_transactions = exceptional_transaction(info_extracted_df)
    high_transactions.name = "high_transactions"
    rtgs_transactions.name = "rtgs_transactions"
    sunday_transactions.name = "sunday_transactions"
    save_to_json(json_dir,[high_transactions,rtgs_transactions,sunday_transactions])
    exceptional_transaction_formatter(high_transactions,rtgs_transactions,sunday_transactions,folder_path,main_report_name)

    # Adding inhouse transactions sheet to the main report
    inhouse_transactions_df,inhouse_transactions_grouped = inhouse_transactions(info_extracted_df)
    inhouse_transactions_df.name = "inhouse_transactions_df"
    inhouse_transactions_grouped.name = "inhouse_transactions_grouped"
    save_to_json(json_dir,[inhouse_transactions_df,inhouse_transactions_grouped])
    inhouse_transactions_formatter(inhouse_transactions_df,inhouse_transactions_grouped,folder_path,main_report_name)

    # Adding salary transactions sheet to the main report
    salary_transactions_df,salary_transactions_grouped = salary_transactions(info_extracted_df)
    salary_transactions_df.name = "salary_transactions_df"
    salary_transactions_grouped.name = "salary_transactions_grouped"
    save_to_json(json_dir,[salary_transactions_df,salary_transactions_grouped])
    salary_transactions_formatter(salary_transactions_df,salary_transactions_grouped,folder_path,main_report_name)

    # Adding loan transactions sheet to the main report
    loan_transactions_df = loan_transactions(info_extracted_df)
    loan_transactions_df.name = "loan_transactions_df"
    save_to_json(json_dir,[loan_transactions_df])
    loan_transactions_formatter(loan_transactions_df,folder_path,main_report_name)

    # Adding interest transactions sheet to the main report
    interest_transactions_df = interest_transactions(info_extracted_df)
    interest_transactions_df.name = "interest_transactions_df"
    save_to_json(json_dir,[interest_transactions_df])
    interest_transactions_formatter(interest_transactions_df,folder_path,main_report_name)

    # Adding transaction return sheet to the main report
    transaction_return_df,transaction_return_grouped = transaction_return(info_extracted_df)
    transaction_return_df.name = "transaction_return_df"
    transaction_return_grouped.name = "transaction_return_grouped"
    save_to_json(json_dir,[transaction_return_df,transaction_return_grouped])
    return_transactions_formatter(transaction_return_df,transaction_return_grouped,folder_path,main_report_name)

    # Adding supplier transactions sheet to the main report
    supplier_transactions_df,supplier_transactions_grouped = supplier_transactions(info_extracted_df)
    supplier_transactions_df.name = "supplier_transactions_df"
    supplier_transactions_grouped.name = "supplier_transactions_grouped"
    save_to_json(json_dir,[supplier_transactions_df,supplier_transactions_grouped])
    supplier_transactions_formatter(supplier_transactions_df,supplier_transactions_grouped,folder_path,main_report_name)

    # Adding customer transactions sheet to the main report
    customer_transactions_df,customer_transactions_grouped = customer_transactions(info_extracted_df)
    customer_transactions_df.name = "customer_transactions_df"
    customer_transactions_grouped.name = "customer_transactions_grouped"
    save_to_json(json_dir,[customer_transactions_df,customer_transactions_grouped])
    customer_transactions_formatter(customer_transactions_df,customer_transactions_grouped,folder_path,main_report_name)

    # Adding charges transactions sheet to the main report
    charges_transactions_df,charges_transactions_grouped = charges_transactions(info_extracted_df)
    charges_transactions_df.name = "charges_transactions_df"
    charges_transactions_grouped.name = "charges_transactions_grouped"
    save_to_json(json_dir,[charges_transactions_df,charges_transactions_grouped])
    charges_transactions_formatter(charges_transactions_df,charges_transactions_grouped,folder_path,main_report_name)

    print("Report Generation Complete")
    return

def report_corrector(info_extracted_df_path,details,party_name,folder_path):
    print("Starting Report Correction")
    # Read the info_extracted_df
    info_extracted_df = pd.read_excel(info_extracted_df_path)
    
    # Redefine the column Cleaned Entity Name
    cleaned_entity_names = set(info_extracted_df['Cleaned Entity Name'].values.tolist())
    grouped_entities = entity_name_grouper(list(cleaned_entity_names))

    for index, row in info_extracted_df.iterrows():
        entity = row['Cleaned Entity Name']
        for key, value in grouped_entities.items():
            if entity in value:
                info_extracted_df.at[index, 'Reconciled Group'] = key    
                reconciled_name = max(value, key=len)
                info_extracted_df.at[index, 'Reconciled Name'] = reconciled_name
                break

    grouped_df = info_extracted_df.groupby('Reconciled Group')
    for name, group in grouped_df:
        # if the group is not Not Found
        if group['Reconciled Name'].iloc[0] != "Not Found" and group['Related Party'].iloc[0] == 'not related party':
            if group.loc[group['Direction'] == 'outward'].shape[0] > 0.7*group.shape[0]:
                flag = 0
                if group.loc[group['Purpose'] == 'others'].shape[0] > 0.7*group.shape[0]:
                    if group.shape[0] > 3:
                            if group.loc[group['Direction'] == 'outward', 'Debit'].std() < 0.1*group.loc[group['Direction'] == 'outward', 'Debit'].mean():
                                group['Date'] = pd.to_datetime(group['Date'], format='%d-%m-%Y')
                                if (group['Date'].diff().dt.days > 25).all() and (group['Date'].diff().dt.days < 35).all():
                                    info_extracted_df.loc[group.index, 'Purpose'] = "salary"
                                    flag = 1
                if flag == 0:
                    info_extracted_df.loc[group.index, 'Purpose'] = "supplier"
            elif group.loc[group['Direction'] == 'inward'].shape[0] > 0.7*group.shape[0]:
                info_extracted_df.loc[group.index, 'Purpose'] = "customer"
    
    # Training the NER model
    try: 
        print("Training NER Model")
        ner_model_trainer(info_extracted_df,iterations=20)   
        print("Training Complete")
    except:
        pass
    from_date, to_date, ifsc_list, bank_list, account_type_list = sheet_details_processor(details)
    
    # Defining a new corrected main report name
    main_report_name = r"\Main_Report_corrected.xlsx"

    # Opening and closing balance
    opening_balance = info_extracted_df['Balance'].iloc[0]
    closing_balance = info_extracted_df['Balance'].iloc[-1]

    # Total months from from_date to to_date
    total_months = (pd.to_datetime(to_date).year - pd.to_datetime(from_date).year)*12 + (pd.to_datetime(to_date).month - pd.to_datetime(from_date).month) + 1   
    
    # Make a workbook object - Main Report Corrected
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Landing Page" # type: ignore
    
    ws['G6'] = "Bank Statement Analysis" # type: ignore
    cell = ws['G6']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, underline='single', color='FFFFFF', size=14)
    cell.fill = PatternFill(start_color='388db1', end_color='388db1', fill_type='solid')
    cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    ws.merge_cells('G6:L6')

    for i in range(7,13):
        for j in range(7,12):
            cell = ws.cell(row=i,column=j)
            cell.fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['G7'] = "Party Name"
    cell = ws['G7']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G7:I7')

    ws['J7'] = party_name
    cell = ws['J7']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J7:L7')

    ws['G8'] = "Start of the Period"
    cell = ws['G8']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G8:I8')

    ws['J8'] = from_date
    cell = ws['J8']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J8:L8')

    ws['G9'] = "End of the Period"
    cell = ws['G9']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G9:I9')

    ws['J9'] = to_date
    cell = ws['J9']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J9:L9')

    ws['G10'] = "Total Months"
    cell = ws['G10']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G10:I10')

    ws['J10'] = total_months
    cell = ws['J10']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('J10:L10')

    ws['G11'] = "Opening Balance"
    cell = ws['G11']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G11:I11')

    ws['J11'] = opening_balance
    cell = ws['J11']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    ws.merge_cells('J11:L11')

    ws['G12'] = "Closing Balance"
    cell = ws['G12']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    ws.merge_cells('G12:I12')

    ws['J12'] = closing_balance
    cell = ws['J12']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    ws.merge_cells('J12:L12')
    ws.sheet_view.showGridLines = False
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value) # type: ignore
            except:
                pass
        adjusted_width = (max_length+5)
        ws.column_dimensions[get_excel_column_name(column-1)].width = adjusted_width
    wb.active = 0
    wb.save(folder_path+main_report_name)

    json_dir = folder_path + r"\json_files"
    try:
        os.makedirs(json_dir,exist_ok=True)
    except:
        pass
    
    # see if statement_details.xlsx is present in the folder
    if os.path.exists(folder_path+r"\statement_details.xlsx"):
        statement_details = pd.read_excel(folder_path+r"\statement_details.xlsx")
        statement_details.name = "statement_details"
        save_to_json(json_dir,[statement_details])
        statement_details_formatter(statement_details,folder_path,main_report_name)

    # Adding graphs to the main report
    info_extracted_df.name = "info_extracted_df"
    save_to_json(json_dir,[info_extracted_df])
    graph_generator(info_extracted_df,folder_path,main_report_name)

    # Adding overview sheet to the main report
    overview_sheet = overview_generator(info_extracted_df)
    overview_sheet.name = "overview_sheet"
    save_to_json(json_dir,[overview_sheet])
    overview_sheet_formatter(overview_sheet,folder_path,main_report_name)

    # Adding summary sheet to the main report
    summary_sheet = summary_of_debits_and_credits(info_extracted_df)
    summary_sheet.name = "summary_sheet"
    save_to_json(json_dir,[summary_sheet])
    summary_sheet_formatter(summary_sheet,folder_path,main_report_name)

    # Adding top credits and debits sheet to the main report
    top_credits,top_debits,top_credits_debits_monthwise = top_credits_and_debits(info_extracted_df)
    top_credits.name = "top_credits"
    top_debits.name = "top_debits"
    top_credits_debits_monthwise.name = "top_credits_debits_monthwise"
    save_to_json(json_dir,[top_credits,top_debits,top_credits_debits_monthwise])
    top_credit_debit_formatter(top_credits,top_debits,top_credits_debits_monthwise,folder_path,main_report_name)

    # Adding exceptional transactions sheet to the main report
    high_transactions,rtgs_transactions,sunday_transactions = exceptional_transaction(info_extracted_df)
    high_transactions.name = "high_transactions"
    rtgs_transactions.name = "rtgs_transactions"
    sunday_transactions.name = "sunday_transactions"
    save_to_json(json_dir,[high_transactions,rtgs_transactions,sunday_transactions])
    exceptional_transaction_formatter(high_transactions,rtgs_transactions,sunday_transactions,folder_path,main_report_name)

    # Adding inhouse transactions sheet to the main report
    inhouse_transactions_df,inhouse_transactions_grouped = inhouse_transactions(info_extracted_df)
    inhouse_transactions_df.name = "inhouse_transactions_df"
    inhouse_transactions_grouped.name = "inhouse_transactions_grouped"
    save_to_json(json_dir,[inhouse_transactions_df,inhouse_transactions_grouped])
    inhouse_transactions_formatter(inhouse_transactions_df,inhouse_transactions_grouped,folder_path,main_report_name)

    # Adding salary transactions sheet to the main report
    salary_transactions_df,salary_transactions_grouped = salary_transactions(info_extracted_df)
    salary_transactions_df.name = "salary_transactions_df"
    salary_transactions_grouped.name = "salary_transactions_grouped"
    save_to_json(json_dir,[salary_transactions_df,salary_transactions_grouped])
    salary_transactions_formatter(salary_transactions_df,salary_transactions_grouped,folder_path,main_report_name)

    # Adding loan transactions sheet to the main report
    loan_transactions_df = loan_transactions(info_extracted_df)
    loan_transactions_df.name = "loan_transactions_df"
    save_to_json(json_dir,[loan_transactions_df])
    loan_transactions_formatter(loan_transactions_df,folder_path,main_report_name)

    # Adding interest transactions sheet to the main report
    interest_transactions_df = interest_transactions(info_extracted_df)
    interest_transactions_df.name = "interest_transactions_df"
    save_to_json(json_dir,[interest_transactions_df])
    interest_transactions_formatter(interest_transactions_df,folder_path,main_report_name)

    # Adding transaction return sheet to the main report
    transaction_return_df,transaction_return_grouped = transaction_return(info_extracted_df)
    transaction_return_df.name = "transaction_return_df"
    transaction_return_grouped.name = "transaction_return_grouped"
    save_to_json(json_dir,[transaction_return_df,transaction_return_grouped])
    return_transactions_formatter(transaction_return_df,transaction_return_grouped,folder_path,main_report_name)

    # Adding supplier transactions sheet to the main report
    supplier_transactions_df,supplier_transactions_grouped = supplier_transactions(info_extracted_df)
    supplier_transactions_df.name = "supplier_transactions_df"
    supplier_transactions_grouped.name = "supplier_transactions_grouped"
    save_to_json(json_dir,[supplier_transactions_df,supplier_transactions_grouped])
    supplier_transactions_formatter(supplier_transactions_df,supplier_transactions_grouped,folder_path,main_report_name)

    # Adding customer transactions sheet to the main report
    customer_transactions_df,customer_transactions_grouped = customer_transactions(info_extracted_df)
    customer_transactions_df.name = "customer_transactions_df"
    customer_transactions_grouped.name = "customer_transactions_grouped"
    save_to_json(json_dir,[customer_transactions_df,customer_transactions_grouped])
    customer_transactions_formatter(customer_transactions_df,customer_transactions_grouped,folder_path,main_report_name)

    # Adding charges transactions sheet to the main report
    charges_transactions_df,charges_transactions_grouped = charges_transactions(info_extracted_df)
    charges_transactions_df.name = "charges_transactions_df"
    charges_transactions_grouped.name = "charges_transactions_grouped"
    save_to_json(json_dir,[charges_transactions_df,charges_transactions_grouped])
    charges_transactions_formatter(charges_transactions_df,charges_transactions_grouped,folder_path,main_report_name)

    print("Report Generation Complete")
    return
